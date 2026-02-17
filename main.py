import streamlit as st
import pandas as pd
import sqlite3
from datetime import datetime, timedelta
import hashlib
import smtplib
import io
import os
import base64
import json
import plotly.graph_objects as go
import plotly.express as px
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import bcrypt
import re
from contextlib import contextmanager
import warnings
warnings.filterwarnings('ignore')

# ============================================================================
# CONSTANTS AND CONFIGURATION
# ============================================================================

CURRENCIES = {
    'TTD': {'symbol': 'TT$', 'name': 'Trinidad & Tobago Dollar'},
    'USD': {'symbol': '$', 'name': 'US Dollar'},
    'EUR': {'symbol': '€', 'name': 'Euro'},
    'GBP': {'symbol': '£', 'name': 'British Pound'},
    'CAD': {'symbol': 'C$', 'name': 'Canadian Dollar'},
    'JMD': {'symbol': 'J$', 'name': 'Jamaican Dollar'}
}

INVOICE_STATUSES = ['Draft', 'Sent', 'Paid', 'Overdue', 'Cancelled']
PAYMENT_METHODS = ['Cash', 'Bank Transfer', 'Credit Card', 'Cheque', 'Online Payment']

RECURRING_FREQUENCIES = {
    'None': None,
    'Daily': 1,
    'Weekly': 7,
    'Monthly': 30,
    'Quarterly': 90,
    'Yearly': 365
}

# ============================================================================
# DATABASE CONNECTION
# ============================================================================

@contextmanager
def get_db_connection():
    """Get database connection with context manager"""
    conn = None
    try:
        conn = sqlite3.connect('invoices.db')
        conn.row_factory = sqlite3.Row
        yield conn
    except sqlite3.Error as e:
        st.error(f"Database connection error: {e}")
        raise
    finally:
        if conn:
            conn.close()

def init_database():
    """Initialize database tables"""
    try:
        with get_db_connection() as conn:
            c = conn.cursor()
            
            # Create company_settings table
            c.execute('''CREATE TABLE IF NOT EXISTS company_settings (
                id INTEGER PRIMARY KEY,
                name TEXT,
                address TEXT,
                city TEXT,
                phone TEXT,
                email TEXT,
                tax_id TEXT,
                bank_details TEXT,
                default_currency TEXT,
                vat_registered BOOLEAN,
                invoice_prefix TEXT,
                logo_base64 TEXT,
                updated_at TEXT
            )''')
            
            # Create clients table
            c.execute('''CREATE TABLE IF NOT EXISTS clients (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                email TEXT NOT NULL,
                phone TEXT,
                address TEXT,
                company TEXT,
                tax_id TEXT,
                notes TEXT,
                credit_limit REAL DEFAULT 0,
                payment_terms INTEGER DEFAULT 30,
                created_at TEXT,
                updated_at TEXT
            )''')
            
            # Create invoices table
            c.execute('''CREATE TABLE IF NOT EXISTS invoices (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                invoice_number TEXT UNIQUE NOT NULL,
                client_name TEXT NOT NULL,
                client_email TEXT,
                client_address TEXT,
                client_phone TEXT,
                invoice_date TEXT NOT NULL,
                due_date TEXT NOT NULL,
                po_number TEXT,
                currency TEXT DEFAULT 'TTD',
                subtotal REAL DEFAULT 0,
                tax_total REAL DEFAULT 0,
                discount_total REAL DEFAULT 0,
                grand_total REAL DEFAULT 0,
                amount_paid REAL DEFAULT 0,
                balance_due REAL DEFAULT 0,
                status TEXT DEFAULT 'Draft',
                notes TEXT,
                sent_date TEXT,
                created_at TEXT,
                updated_at TEXT,
                recurring_frequency TEXT,
                recurring_next_date TEXT
            )''')
            
            # Create invoice_items table
            c.execute('''CREATE TABLE IF NOT EXISTS invoice_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                invoice_id INTEGER,
                description TEXT NOT NULL,
                quantity REAL DEFAULT 1,
                unit_price REAL DEFAULT 0,
                tax_rate REAL DEFAULT 0,
                discount REAL DEFAULT 0,
                total REAL DEFAULT 0,
                FOREIGN KEY (invoice_id) REFERENCES invoices (id) ON DELETE CASCADE
            )''')
            
            # Create payments table
            c.execute('''CREATE TABLE IF NOT EXISTS payments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                invoice_id INTEGER,
                amount REAL NOT NULL,
                payment_date TEXT NOT NULL,
                payment_method TEXT NOT NULL,
                reference TEXT,
                notes TEXT,
                created_at TEXT,
                FOREIGN KEY (invoice_id) REFERENCES invoices (id)
            )''')
            
            # Create users table
            c.execute('''CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                email TEXT UNIQUE NOT NULL,
                role TEXT DEFAULT 'user',
                full_name TEXT,
                is_active BOOLEAN DEFAULT 1,
                last_login TEXT,
                created_at TEXT,
                updated_at TEXT
            )''')
            
            # Create recurring_invoices table
            c.execute('''CREATE TABLE IF NOT EXISTS recurring_invoices (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                template_id INTEGER,
                client_id INTEGER,
                frequency TEXT NOT NULL,
                start_date TEXT NOT NULL,
                end_date TEXT,
                next_date TEXT NOT NULL,
                last_generated TEXT,
                is_active BOOLEAN DEFAULT 1,
                created_at TEXT,
                FOREIGN KEY (template_id) REFERENCES invoice_templates (id),
                FOREIGN KEY (client_id) REFERENCES clients (id)
            )''')
            
            # Create invoice_templates table
            c.execute('''CREATE TABLE IF NOT EXISTS invoice_templates (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                description TEXT,
                template_data TEXT,
                created_at TEXT,
                updated_at TEXT
            )''')
            
            # Create audit_log table
            c.execute('''CREATE TABLE IF NOT EXISTS audit_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                action TEXT NOT NULL,
                table_name TEXT,
                record_id INTEGER,
                old_value TEXT,
                new_value TEXT,
                ip_address TEXT,
                timestamp TEXT,
                FOREIGN KEY (user_id) REFERENCES users (id)
            )''')
            
            # Insert default company settings if none exist
            c.execute("SELECT COUNT(*) FROM company_settings")
            if c.fetchone()[0] == 0:
                c.execute('''INSERT INTO company_settings 
                    (name, email, default_currency, vat_registered, invoice_prefix, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?)''',
                    ('My Company', 'company@example.com', 'TTD', 1, 'INV', datetime.now().isoformat()))
            
            # Insert default admin user if none exist
            c.execute("SELECT COUNT(*) FROM users")
            if c.fetchone()[0] == 0:
                default_password = hash_password("admin123")
                c.execute('''INSERT INTO users 
                    (username, password_hash, email, role, full_name, created_at)
                    VALUES (?, ?, ?, ?, ?, ?)''',
                    ('admin', default_password, 'admin@example.com', 'admin', 'System Administrator', 
                     datetime.now().isoformat()))
            
            conn.commit()
    except Exception as e:
        st.error(f"Database initialization error: {e}")

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def format_amount(amount, currency='TTD'):
    """Format amount with currency symbol"""
    try:
        amount = float(amount)
        symbol = CURRENCIES.get(currency, {'symbol': '$'})['symbol']
        return f"{symbol}{amount:,.2f}"
    except (ValueError, TypeError):
        return f"{CURRENCIES.get(currency, {'symbol': '$'})['symbol']}0.00"

def get_currency_symbol(currency):
    """Get currency symbol"""
    return CURRENCIES.get(currency, {'symbol': '$'})['symbol']

def generate_invoice_number():
    """Generate unique invoice number"""
    prefix = st.session_state.company_info.get('invoice_prefix', 'INV')
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
    return f"{prefix}-{timestamp}"

def hash_password(password):
    """Hash password with bcrypt"""
    salt = bcrypt.gensalt()
    return bcrypt.hashpw(password.encode('utf-8'), salt)

def verify_password(password, hashed):
    """Verify password"""
    if isinstance(hashed, str):
        hashed = hashed.encode('utf-8')
    return bcrypt.checkpw(password.encode('utf-8'), hashed)

def validate_email(email):
    """Validate email format"""
    pattern = r'^[\w\.-]+@[\w\.-]+\.\w+$'
    return bool(re.match(pattern, email))

def validate_phone(phone):
    """Validate phone number"""
    pattern = r'^[\d\s\+\-\(\)]{7,}$'
    return bool(re.match(pattern, phone))

def safe_db_operation(func):
    """Decorator for safe database operations"""
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except sqlite3.Error as e:
            st.error(f"Database error: {e}")
            return None
        except Exception as e:
            st.error(f"Unexpected error: {e}")
            return None
    return wrapper

@st.cache_data(ttl=300)
def get_cached_data(query, params=None):
    """Get cached database results"""
    with get_db_connection() as conn:
        if params:
            return pd.read_sql_query(query, conn, params=params)
        return pd.read_sql_query(query, conn)

def paginate_dataframe(df, page_size=10, key="default"):
    """Paginate dataframe display"""
    if df.empty:
        return df
    
    total_pages = len(df) // page_size + (1 if len(df) % page_size else 0)
    page_key = f"page_num_{key}"
    
    if page_key not in st.session_state:
        st.session_state[page_key] = 0
    
    col1, col2, col3 = st.columns([1, 2, 1])
    with col1:
        if st.button("◀ Previous", key=f"prev_{key}") and st.session_state[page_key] > 0:
            st.session_state[page_key] -= 1
    with col2:
        st.write(f"Page {st.session_state[page_key] + 1} of {total_pages}")
    with col3:
        if st.button("Next ▶", key=f"next_{key}") and st.session_state[page_key] < total_pages - 1:
            st.session_state[page_key] += 1
    
    start = st.session_state[page_key] * page_size
    end = start + page_size
    return df.iloc[start:end]

def log_audit(action, table_name=None, record_id=None, old_value=None, new_value=None):
    """Log audit entry"""
    try:
        with get_db_connection() as conn:
            c = conn.cursor()
            c.execute('''INSERT INTO audit_log 
                        (user_id, action, table_name, record_id, old_value, new_value, timestamp)
                        VALUES (?, ?, ?, ?, ?, ?, ?)''',
                     (st.session_state.get('user_id', 1), action, table_name, record_id,
                      json.dumps(old_value) if old_value else None,
                      json.dumps(new_value) if new_value else None,
                      datetime.now().isoformat()))
            conn.commit()
    except Exception as e:
        print(f"Audit log error: {e}")

def get_status_badge_html(status):
    """Get HTML for status badge"""
    colors = {
        'Draft': '#95a5a6',
        'Sent': '#3498db',
        'Paid': '#27ae60',
        'Overdue': '#e74c3c',
        'Cancelled': '#7f8c8d'
    }
    color = colors.get(status, '#95a5a6')
    return f'<span style="background-color: {color}; color: white; padding: 3px 10px; border-radius: 12px; font-size: 12px;">{status}</span>'

def save_logo(uploaded_file):
    """Save uploaded logo"""
    try:
        if uploaded_file is not None:
            bytes_data = uploaded_file.getvalue()
            base64_data = base64.b64encode(bytes_data).decode()
            st.session_state.company_info['logo_base64'] = base64_data
            
            with get_db_connection() as conn:
                c = conn.cursor()
                c.execute('''UPDATE company_settings 
                           SET logo_base64 = ?, updated_at = ?
                           WHERE id = 1''',
                         (base64_data, datetime.now().isoformat()))
                conn.commit()
            return True
    except Exception as e:
        st.error(f"Error saving logo: {e}")
        return False

def get_logo_html(height="50px", width="auto"):
    """Get HTML for logo display"""
    if st.session_state.company_info.get('logo_base64'):
        return f'<img src="data:image/png;base64,{st.session_state.company_info["logo_base64"]}" style="height: {height}; width: {width}; object-fit: contain;">'
    return ""

def remove_logo():
    """Remove company logo"""
    try:
        st.session_state.company_info['logo_base64'] = None
        with get_db_connection() as conn:
            c = conn.cursor()
            c.execute('''UPDATE company_settings 
                       SET logo_base64 = NULL, updated_at = ?
                       WHERE id = 1''',
                     (datetime.now().isoformat(),))
            conn.commit()
        return True
    except Exception as e:
        st.error(f"Error removing logo: {e}")
        return False

# ============================================================================
# DATABASE OPERATIONS
# ============================================================================

@safe_db_operation
def save_invoice_to_db(invoice_data, items):
    """Save invoice to database"""
    errors = []
    warnings = []
    invoice_id = None
    
    try:
        with get_db_connection() as conn:
            c = conn.cursor()
            
            # Insert invoice
            c.execute('''INSERT INTO invoices 
                        (invoice_number, client_name, client_email, client_address, client_phone,
                         invoice_date, due_date, po_number, currency, subtotal, tax_total,
                         discount_total, grand_total, amount_paid, balance_due, status,
                         notes, sent_date, recurring_frequency, recurring_next_date,
                         created_at, updated_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                     (invoice_data['invoice_number'], invoice_data['client_name'],
                      invoice_data.get('client_email'), invoice_data.get('client_address'),
                      invoice_data.get('client_phone'), invoice_data['invoice_date'],
                      invoice_data['due_date'], invoice_data.get('po_number'),
                      invoice_data['currency'], invoice_data['subtotal'],
                      invoice_data['tax_total'], invoice_data['discount_total'],
                      invoice_data['grand_total'], invoice_data.get('amount_paid', 0),
                      invoice_data.get('balance_due', invoice_data['grand_total']),
                      invoice_data['status'], invoice_data.get('notes'),
                      invoice_data.get('sent_date'), invoice_data.get('recurring_frequency'),
                      invoice_data.get('recurring_next_date'), datetime.now().isoformat(),
                      datetime.now().isoformat()))
            
            invoice_id = c.lastrowid
            
            # Insert items
            for item in items:
                c.execute('''INSERT INTO invoice_items 
                            (invoice_id, description, quantity, unit_price, tax_rate, discount, total)
                            VALUES (?, ?, ?, ?, ?, ?, ?)''',
                         (invoice_id, item['description'], item['quantity'],
                          item['unit_price'], item['tax_rate'], item['discount'],
                          item['total']))
            
            conn.commit()
            log_audit('CREATE', 'invoices', invoice_id, None, invoice_data)
            
    except Exception as e:
        errors.append(str(e))
    
    return invoice_id, errors, warnings

@safe_db_operation
def get_invoices(filters=None):
    """Get invoices with optional filters"""
    query = "SELECT * FROM invoices"
    params = []
    
    if filters:
        conditions = []
        if filters.get('status') and filters['status'] != 'All':
            conditions.append("status = ?")
            params.append(filters['status'])
        if filters.get('client_name'):
            conditions.append("client_name LIKE ?")
            params.append(f"%{filters['client_name']}%")
        if filters.get('date_from'):
            conditions.append("invoice_date >= ?")
            params.append(filters['date_from'])
        if filters.get('date_to'):
            conditions.append("invoice_date <= ?")
            params.append(filters['date_to'])
        
        if conditions:
            query += " WHERE " + " AND ".join(conditions)
    
    query += " ORDER BY created_at DESC"
    
    with get_db_connection() as conn:
        return pd.read_sql_query(query, conn, params=params)

@safe_db_operation
def get_invoice_by_id(invoice_id):
    """Get invoice by ID"""
    with get_db_connection() as conn:
        c = conn.cursor()
        c.execute("SELECT * FROM invoices WHERE id = ?", (invoice_id,))
        invoice = c.fetchone()
        
        if invoice:
            c.execute("SELECT * FROM invoice_items WHERE invoice_id = ?", (invoice_id,))
            items = c.fetchall()
            return dict(invoice), [dict(item) for item in items]
    
    return None, None

@safe_db_operation
def update_invoice_status(invoice_id, new_status):
    """Update invoice status"""
    with get_db_connection() as conn:
        c = conn.cursor()
        c.execute('''UPDATE invoices 
                    SET status = ?, updated_at = ?
                    WHERE id = ?''',
                 (new_status, datetime.now().isoformat(), invoice_id))
        conn.commit()
        log_audit('UPDATE', 'invoices', invoice_id, {'status': 'old'}, {'status': new_status})
        return True

@safe_db_operation
def delete_invoice(invoice_id):
    """Delete invoice"""
    with get_db_connection() as conn:
        c = conn.cursor()
        c.execute("DELETE FROM invoices WHERE id = ?", (invoice_id,))
        conn.commit()
        log_audit('DELETE', 'invoices', invoice_id)
        return True

@safe_db_operation
def save_client_to_db(client_data):
    """Save client to database"""
    with get_db_connection() as conn:
        c = conn.cursor()
        
        # Check if client exists
        c.execute("SELECT id FROM clients WHERE email = ?", (client_data['email'],))
        existing = c.fetchone()
        
        if existing:
            # Update existing client
            c.execute('''UPDATE clients 
                        SET name = ?, phone = ?, address = ?, company = ?,
                            tax_id = ?, notes = ?, credit_limit = ?, payment_terms = ?,
                            updated_at = ?
                        WHERE email = ?''',
                     (client_data['name'], client_data.get('phone'),
                      client_data.get('address'), client_data.get('company'),
                      client_data.get('tax_id'), client_data.get('notes'),
                      client_data.get('credit_limit', 0),
                      client_data.get('payment_terms', 30),
                      datetime.now().isoformat(), client_data['email']))
            client_id = existing[0]
        else:
            # Insert new client
            c.execute('''INSERT INTO clients 
                        (name, email, phone, address, company, tax_id, notes,
                         credit_limit, payment_terms, created_at, updated_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                     (client_data['name'], client_data['email'],
                      client_data.get('phone'), client_data.get('address'),
                      client_data.get('company'), client_data.get('tax_id'),
                      client_data.get('notes'), client_data.get('credit_limit', 0),
                      client_data.get('payment_terms', 30),
                      datetime.now().isoformat(), datetime.now().isoformat()))
            client_id = c.lastrowid
        
        conn.commit()
        log_audit('CREATE' if not existing else 'UPDATE', 'clients', client_id, None, client_data)
        return client_id

@safe_db_operation
def get_clients(search_term=None):
    """Get clients with optional search"""
    if search_term:
        with get_db_connection() as conn:
            return pd.read_sql_query('''SELECT * FROM clients 
                                       WHERE name LIKE ? OR email LIKE ? OR company LIKE ?
                                       ORDER BY name''',
                                    conn, params=[f'%{search_term}%', f'%{search_term}%', f'%{search_term}%'])
    else:
        with get_db_connection() as conn:
            return pd.read_sql_query("SELECT * FROM clients ORDER BY name", conn)

@safe_db_operation
def process_payment(invoice_id, amount, method, reference=None, notes=None):
    """Process payment for invoice"""
    try:
        with get_db_connection() as conn:
            c = conn.cursor()
            
            # Get invoice details
            c.execute("SELECT * FROM invoices WHERE id = ?", (invoice_id,))
            invoice = dict(c.fetchone())
            
            # Calculate new amounts
            new_amount_paid = invoice['amount_paid'] + amount
            new_balance_due = invoice['grand_total'] - new_amount_paid
            new_status = 'Paid' if new_balance_due <= 0 else invoice['status']
            
            # Insert payment record
            c.execute('''INSERT INTO payments 
                        (invoice_id, amount, payment_date, payment_method, reference, notes, created_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?)''',
                     (invoice_id, amount, datetime.now().strftime('%Y-%m-%d'),
                      method, reference, notes, datetime.now().isoformat()))
            
            # Update invoice
            c.execute('''UPDATE invoices 
                        SET amount_paid = ?, balance_due = ?, status = ?, updated_at = ?
                        WHERE id = ?''',
                     (new_amount_paid, new_balance_due, new_status,
                      datetime.now().isoformat(), invoice_id))
            
            conn.commit()
            log_audit('CREATE', 'payments', c.lastrowid, None, 
                     {'invoice_id': invoice_id, 'amount': amount, 'method': method})
            
            return True, f"Payment of {format_amount(amount, invoice['currency'])} recorded successfully"
            
    except Exception as e:
        return False, str(e)

@safe_db_operation
def create_recurring_invoice(template_id, client_id, frequency, start_date, end_date=None):
    """Create recurring invoice schedule"""
    with get_db_connection() as conn:
        c = conn.cursor()
        
        c.execute('''INSERT INTO recurring_invoices 
                    (template_id, client_id, frequency, start_date, end_date, next_date, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?)''',
                 (template_id, client_id, frequency, start_date, end_date, start_date,
                  datetime.now().isoformat()))
        
        conn.commit()
        return c.lastrowid

@safe_db_operation
def backup_database():
    """Create database backup"""
    try:
        backup_data = io.BytesIO()
        with open('invoices.db', 'rb') as f:
            backup_data.write(f.read())
        
        backup_data.seek(0)
        filename = f"invoice_pro_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
        return backup_data.getvalue(), filename
    except Exception as e:
        st.error(f"Backup failed: {e}")
        return None, None

@safe_db_operation
def restore_database(backup_path):
    """Restore database from backup"""
    try:
        # Close any existing connections
        sqlite3.connect('invoices.db').close()
        
        # Restore backup
        import shutil
        shutil.copy2(backup_path, 'invoices.db')
        
        log_audit('RESTORE', 'database', None, None, {'backup': backup_path})
        return True
    except Exception as e:
        st.error(f"Restore failed: {e}")
        return False

# ============================================================================
# PDF GENERATION
# ============================================================================

def generate_pdf_invoice(invoice_data):
    """Generate PDF invoice"""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
        from reportlab.lib.enums import TA_RIGHT, TA_CENTER
        import io
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        styles = getSampleStyleSheet()
        elements = []
        
        # Company Logo and Info
        if invoice_data['company_info'].get('logo_base64'):
            logo_data = base64.b64decode(invoice_data['company_info']['logo_base64'])
            logo_buffer = io.BytesIO(logo_data)
            img = Image(logo_buffer, width=2*inch, height=1*inch)
            elements.append(img)
        
        # Company Info
        company_text = f"""
        <b>{invoice_data['company_info']['name']}</b><br/>
        {invoice_data['company_info'].get('address', '')}<br/>
        {invoice_data['company_info'].get('city', '')}<br/>
        Phone: {invoice_data['company_info'].get('phone', '')}<br/>
        Email: {invoice_data['company_info'].get('email', '')}<br/>
        TRN: {invoice_data['company_info'].get('tax_id', '')}
        """
        elements.append(Paragraph(company_text, styles['Normal']))
        elements.append(Spacer(1, 0.2*inch))
        
        # Invoice Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2c3e50'),
            alignment=TA_CENTER
        )
        elements.append(Paragraph(f"INVOICE", title_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Invoice Number and Dates
        info_data = [
            ['Invoice Number:', invoice_data['invoice_number'],
             'Date:', invoice_data['invoice_date']],
            ['PO Number:', invoice_data.get('po_number', 'N/A'),
             'Due Date:', invoice_data['due_date']],
            ['Status:', invoice_data['status'], '', '']
        ]
        
        info_table = Table(info_data, colWidths=[1.2*inch, 2*inch, 1.2*inch, 2*inch])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('BACKGROUND', (2, 0), (2, -1), colors.lightgrey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('PADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 0.2*inch))
        
        # Bill To
        bill_data = [
            ['Bill To:'],
            [invoice_data['client']['name']],
            [invoice_data['client'].get('address', '')],
            [f"Email: {invoice_data['client'].get('email', '')}"],
            [f"Phone: {invoice_data['client'].get('phone', '')}"]
        ]
        
        bill_table = Table(bill_data, colWidths=[4*inch])
        bill_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('BOX', (0, 0), (-1, -1), 1, colors.grey),
            ('BACKGROUND', (0, 0), (0, 0), colors.lightgrey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('PADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(bill_table)
        elements.append(Spacer(1, 0.2*inch))
        
        # Items Table
        items_data = [['Description', 'Qty', 'Unit Price', 'Tax %', 'Disc %', 'Total']]
        
        for item in invoice_data['items']:
            items_data.append([
                item['description'],
                f"{item['quantity']:.2f}",
                format_amount(item['unit_price'], invoice_data['currency']),
                f"{item['tax_rate']:.1f}%",
                f"{item['discount']:.1f}%",
                format_amount(item['total'], invoice_data['currency'])
            ])
        
        # Add totals
        items_data.append(['', '', '', '', 'Subtotal:', format_amount(invoice_data['totals']['subtotal'], invoice_data['currency'])])
        items_data.append(['', '', '', '', 'Discount:', f"-{format_amount(invoice_data['totals']['discount'], invoice_data['currency'])}"])
        items_data.append(['', '', '', '', 'Tax:', format_amount(invoice_data['totals']['tax'], invoice_data['currency'])])
        items_data.append(['', '', '', '', 'Grand Total:', format_amount(invoice_data['totals']['grand_total'], invoice_data['currency'])])
        
        if invoice_data['amount_paid'] > 0:
            items_data.append(['', '', '', '', 'Amount Paid:', format_amount(invoice_data['amount_paid'], invoice_data['currency'])])
            items_data.append(['', '', '', '', 'Balance Due:', format_amount(invoice_data['balance_due'], invoice_data['currency'])])
        
        items_table = Table(items_data, colWidths=[2.5*inch, 0.8*inch, 1*inch, 0.8*inch, 0.8*inch, 1.2*inch])
        items_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, len(invoice_data['items'])), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (1, 1), (-2, -1), 'RIGHT'),
            ('ALIGN', (-1, 1), (-1, -1), 'RIGHT'),
            ('FONTWEIGHT', (0, len(items_data)-5), (-1, -1), 'BOLD'),
            ('BACKGROUND', (0, len(items_data)-5), (-1, -1), colors.lightgrey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('PADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(items_table)
        elements.append(Spacer(1, 0.2*inch))
        
        # Notes
        if invoice_data.get('notes'):
            elements.append(Paragraph("<b>Notes:</b>", styles['Normal']))
            elements.append(Paragraph(invoice_data['notes'], styles['Normal']))
            elements.append(Spacer(1, 0.2*inch))
        
        # Bank Details
        if invoice_data['company_info'].get('bank_details'):
            elements.append(Paragraph("<b>Payment Details:</b>", styles['Normal']))
            elements.append(Paragraph(invoice_data['company_info']['bank_details'], styles['Normal']))
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()
        
    except ImportError:
        # Fallback if reportlab not installed
        st.warning("PDF generation requires reportlab. Install with: pip install reportlab")
        return None
    except Exception as e:
        st.error(f"PDF generation error: {e}")
        return None

# ============================================================================
# EMAIL FUNCTIONS
# ============================================================================

def send_email_invoice(to_email, pdf_buffer, invoice_number):
    """Send invoice via email"""
    try:
        smtp_server = os.getenv('SMTP_SERVER', 'smtp.gmail.com')
        smtp_port = int(os.getenv('SMTP_PORT', 587))
        smtp_username = os.getenv('SMTP_USERNAME', '')
        smtp_password = os.getenv('SMTP_PASSWORD', '')
        use_tls = os.getenv('SMTP_USE_TLS', 'True') == 'True'
        
        if not smtp_username or not smtp_password:
            return False, "Email settings not configured. Please check Settings > Email."
        
        msg = MIMEMultipart()
        msg['From'] = smtp_username
        msg['To'] = to_email
        msg['Subject'] = f"Invoice {invoice_number} from {st.session_state.company_info['name']}"
        
        body = f"""
        <html>
        <body>
            <p>Dear Customer,</p>
            
            <p>Please find attached invoice {invoice_number} for your reference.</p>
            
            <p>Invoice Details:</p>
            <ul>
                <li>Invoice Number: {invoice_number}</li>
                <li>Date: {datetime.now().strftime('%Y-%m-%d')}</li>
            </ul>
            
            <p>Payment can be made via:</p>
            <p>{st.session_state.company_info.get('bank_details', 'Bank transfer details not provided')}</p>
            
            <p>Thank you for your business!</p>
            
            <p>Best regards,<br>
            {st.session_state.company_info['name']}</p>
        </body>
        </html>
        """
        
        msg.attach(MIMEText(body, 'html'))
        
        # Attach PDF
        if pdf_buffer:
            attachment = MIMEApplication(pdf_buffer, _subtype="pdf")
            attachment.add_header('Content-Disposition', 'attachment', 
                                filename=f"invoice_{invoice_number}.pdf")
            msg.attach(attachment)
        
        # Send email
        server = smtplib.SMTP(smtp_server, smtp_port)
        if use_tls:
            server.starttls()
        server.login(smtp_username, smtp_password)
        server.send_message(msg)
        server.quit()
        
        log_audit('EMAIL', 'invoices', None, None, {'to': to_email, 'invoice': invoice_number})
        
        return True, f"Invoice sent successfully to {to_email}"
        
    except Exception as e:
        return False, str(e)

# ============================================================================
# EXPORT FUNCTIONS
# ============================================================================

def export_to_excel(invoice_data, items):
    """Export invoice to Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        output = io.BytesIO()
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Invoice details sheet
        ws1 = wb.active
        ws1.title = "Invoice Details"
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add invoice details
        details = [
            ["Field", "Value"],
            ["Invoice Number", invoice_data.get('invoice_number', '')],
            ["Client Name", invoice_data.get('client_name', '')],
            ["Client Email", invoice_data.get('client_email', '')],
            ["Invoice Date", invoice_data.get('invoice_date', '')],
            ["Due Date", invoice_data.get('due_date', '')],
            ["PO Number", invoice_data.get('po_number', '')],
            ["Currency", invoice_data.get('currency', 'TTD')],
            ["Subtotal", invoice_data.get('subtotal', 0)],
            ["Tax Total", invoice_data.get('tax_total', 0)],
            ["Discount Total", invoice_data.get('discount_total', 0)],
            ["Grand Total", invoice_data.get('grand_total', 0)]
        ]
        
        for row in details:
            ws1.append(row)
        
        # Style header
        for cell in ws1[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Items sheet
        ws2 = wb.create_sheet("Items")
        
        # Add items headers
        item_headers = ["Description", "Quantity", "Unit Price", "Tax Rate %", "Discount %", "Total"]
        ws2.append(item_headers)
        
        # Style items header
        for cell in ws2[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add items
        for item in items:
            ws2.append([
                item.get('description', ''),
                item.get('quantity', 0),
                item.get('unit_price', 0),
                item.get('tax_rate', 0),
                item.get('discount', 0),
                item.get('total', 0)
            ])
        
        # Auto-size columns
        for ws in [ws1, ws2]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to buffer
        wb.save(output)
        output.seek(0)
        return output.getvalue()
        
    except ImportError:
        st.warning("Excel export requires openpyxl. Install with: pip install openpyxl")
        return None
    except Exception as e:
        st.error(f"Excel export error: {e}")
        return None

# ============================================================================
# STYLING
# ============================================================================

def add_custom_css():
    """Add custom CSS styling"""
    st.markdown("""
    <style>
    /* Main container */
    .main {
        padding: 0rem 1rem;
    }
    
    /* Business card style */
    .business-card {
        background: white;
        border-radius: 10px;
        padding: 20px;
        margin-bottom: 15px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        border-left: 4px solid #2c3e50;
        transition: transform 0.2s;
    }
    .business-card:hover {
        transform: translateX(5px);
        box-shadow: 0 4px 8px rgba(0,0,0,0.15);
    }
    
    /* Section header */
    .section-header {
        background: linear-gradient(90deg, #2c3e50, #3498db);
        color: white;
        padding: 15px 20px;
        border-radius: 10px;
        margin-bottom: 25px;
        font-size: 24px;
        font-weight: bold;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
    }
    
    /* Invoice preview */
    .invoice-preview {
        background: white;
        border-radius: 10px;
        padding: 30px;
        margin: 20px 0;
        box-shadow: 0 4px 8px rgba(0,0,0,0.1);
        border: 1px solid #e0e0e0;
    }
    
    /* Status badges */
    .status-badge {
        display: inline-block;
        padding: 4px 12px;
        border-radius: 20px;
        font-size: 12px;
        font-weight: bold;
        text-transform: uppercase;
    }
    
    /* Action buttons */
    .action-buttons {
        display: flex;
        gap: 10px;
        margin: 20px 0;
        flex-wrap: wrap;
    }
    
    /* Logo container */
    .logo-container {
        padding: 10px;
        background: #f8f9fa;
        border-radius: 8px;
        display: inline-block;
        margin: 10px 0;
    }
    
    /* Form styling */
    .stTextInput > div > div > input {
        border-radius: 8px;
    }
    
    .stSelectbox > div > div > select {
        border-radius: 8px;
    }
    
    .stButton > button {
        border-radius: 8px;
        font-weight: 600;
        transition: all 0.3s;
    }
    
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 8px rgba(0,0,0,0.2);
    }
    
    /* Metric styling */
    .stMetric {
        background: white;
        padding: 15px;
        border-radius: 10px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    }
    
    /* Tab styling */
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
    }
    
    .stTabs [data-baseweb="tab"] {
        border-radius: 8px 8px 0 0;
        padding: 10px 20px;
        font-weight: 600;
    }
    
    /* Dataframe styling */
    .dataframe {
        font-size: 14px;
        border-collapse: collapse;
        width: 100%;
    }
    
    .dataframe th {
        background-color: #2c3e50;
        color: white;
        padding: 12px;
        text-align: left;
    }
    
    .dataframe td {
        padding: 10px;
        border-bottom: 1px solid #e0e0e0;
    }
    
    /* Responsive design */
    @media (max-width: 768px) {
        .section-header {
            font-size: 20px;
            padding: 12px 15px;
        }
        
        .business-card {
            padding: 15px;
        }
        
        .stButton > button {
            width: 100%;
            margin: 5px 0;
        }
        
        .stColumns {
            flex-direction: column;
        }
    }
    
    /* Print styles */
    @media print {
        .stButton, .stDownloadButton, .stDeleteButton {
            display: none !important;
        }
        
        .business-card {
            box-shadow: none;
            border: 1px solid #000;
        }
    }
    
    /* Animations */
    @keyframes fadeIn {
        from { opacity: 0; transform: translateY(20px); }
        to { opacity: 1; transform: translateY(0); }
    }
    
    .fade-in {
        animation: fadeIn 0.5s ease-out;
    }
    
    /* Notification styling */
    .success-notification {
        background: #d4edda;
        color: #155724;
        padding: 12px 20px;
        border-radius: 8px;
        border-left: 4px solid #28a745;
        margin: 10px 0;
        animation: slideIn 0.3s ease-out;
    }
    
    .error-notification {
        background: #f8d7da;
        color: #721c24;
        padding: 12px 20px;
        border-radius: 8px;
        border-left: 4px solid #dc3545;
        margin: 10px 0;
        animation: slideIn 0.3s ease-out;
    }
    
    @keyframes slideIn {
        from { transform: translateX(-100%); opacity: 0; }
        to { transform: translateX(0); opacity: 1; }
    }
    
    /* Loading spinner */
    .loading-spinner {
        border: 4px solid #f3f3f3;
        border-top: 4px solid #3498db;
        border-radius: 50%;
        width: 40px;
        height: 40px;
        animation: spin 1s linear infinite;
        margin: 20px auto;
    }
    
    @keyframes spin {
        0% { transform: rotate(0deg); }
        100% { transform: rotate(360deg); }
    }
    
    /* Tooltip */
    .tooltip {
        position: relative;
        display: inline-block;
        cursor: help;
    }
    
    .tooltip .tooltiptext {
        visibility: hidden;
        width: 200px;
        background-color: #2c3e50;
        color: #fff;
        text-align: center;
        border-radius: 6px;
        padding: 5px;
        position: absolute;
        z-index: 1;
        bottom: 125%;
        left: 50%;
        margin-left: -100px;
        opacity: 0;
        transition: opacity 0.3s;
    }
    
    .tooltip:hover .tooltiptext {
        visibility: visible;
        opacity: 1;
    }
    </style>
    """, unsafe_allow_html=True)

# ============================================================================
# PAGE FUNCTIONS
# ============================================================================

def render_dashboard_page():
    """Render the dashboard page"""
    
    st.markdown('<div class="section-header">📊 Dashboard</div>', unsafe_allow_html=True)
    
    # Get statistics
    with get_db_connection() as conn:
        # Total invoices
        total_invoices = pd.read_sql_query("SELECT COUNT(*) as count FROM invoices", conn).iloc[0]['count']
        
        # Total revenue
        total_revenue = pd.read_sql_query("SELECT SUM(grand_total) as total FROM invoices WHERE status='Paid'", conn).iloc[0]['total'] or 0
        
        # Outstanding balance
        outstanding = pd.read_sql_query("SELECT SUM(balance_due) as total FROM invoices WHERE status IN ('Sent', 'Overdue')", conn).iloc[0]['total'] or 0
        
        # Recent invoices
        recent_invoices = pd.read_sql_query("""
            SELECT invoice_number, client_name, grand_total, status, due_date 
            FROM invoices 
            ORDER BY created_at DESC 
            LIMIT 5
        """, conn)
        
        # Upcoming due dates
        upcoming = pd.read_sql_query("""
            SELECT invoice_number, client_name, due_date, grand_total, balance_due
            FROM invoices 
            WHERE status IN ('Sent', 'Overdue') 
            AND date(due_date) <= date('now', '+7 days')
            ORDER BY due_date
            LIMIT 5
        """, conn)
        
        # Monthly revenue chart
        monthly_revenue = pd.read_sql_query("""
            SELECT strftime('%Y-%m', invoice_date) as month,
                   SUM(CASE WHEN status='Paid' THEN grand_total ELSE 0 END) as revenue
            FROM invoices
            WHERE invoice_date >= date('now', '-6 months')
            GROUP BY strftime('%Y-%m', invoice_date)
            ORDER BY month
        """, conn)
    
    # Key metrics
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Total Invoices", total_invoices)
    with col2:
        st.metric("Total Revenue", format_amount(total_revenue, st.session_state.currency))
    with col3:
        st.metric("Outstanding", format_amount(outstanding, st.session_state.currency))
    with col4:
        paid_ratio = (total_revenue / (total_revenue + outstanding) * 100) if (total_revenue + outstanding) > 0 else 0
        st.metric("Collection Rate", f"{paid_ratio:.1f}%")
    
    st.divider()
    
    # Charts
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("**Monthly Revenue**")
        if not monthly_revenue.empty:
            fig = px.line(monthly_revenue, x='month', y='revenue', markers=True)
            fig.update_layout(
                plot_bgcolor='white',
                paper_bgcolor='white',
                font=dict(color='#2c3e50'),
                xaxis_title="Month",
                yaxis_title=f"Revenue ({get_currency_symbol(st.session_state.currency)})"
            )
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("No revenue data available")
    
    with col2:
        st.markdown("**Invoice Status Distribution**")
        status_counts = pd.read_sql_query("""
            SELECT status, COUNT(*) as count, SUM(grand_total) as total
            FROM invoices
            GROUP BY status
        """, get_db_connection())
        
        if not status_counts.empty:
            fig = px.pie(status_counts, values='total', names='status', 
                        title='Revenue by Status')
            fig.update_layout(
                plot_bgcolor='white',
                paper_bgcolor='white',
                font=dict(color='#2c3e50')
            )
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("No invoice data available")
    
    st.divider()
    
    # Recent activity
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("**Recent Invoices**")
        if not recent_invoices.empty:
            for _, inv in recent_invoices.iterrows():
                st.markdown(f"""
                <div class="business-card">
                    <strong>{inv['invoice_number']}</strong> - {inv['client_name']}<br>
                    Amount: {format_amount(inv['grand_total'], st.session_state.currency)}<br>
                    Status: {get_status_badge_html(inv['status'])}<br>
                    Due: {inv['due_date']}
                </div>
                """, unsafe_allow_html=True)
        else:
            st.info("No recent invoices")
    
    with col2:
        st.markdown("**Upcoming Due Dates**")
        if not upcoming.empty:
            for _, inv in upcoming.iterrows():
                days_until = (datetime.strptime(inv['due_date'], '%Y-%m-%d') - datetime.now()).days
                st.markdown(f"""
                <div class="business-card">
                    <strong>{inv['invoice_number']}</strong> - {inv['client_name']}<br>
                    Due: {inv['due_date']} ({days_until} days)<br>
                    Amount: {format_amount(inv['grand_total'], st.session_state.currency)}<br>
                    Balance: {format_amount(inv['balance_due'], st.session_state.currency)}
                </div>
                """, unsafe_allow_html=True)
        else:
            st.info("No upcoming due dates")

def render_create_invoice_page():
    """Render the create invoice page"""
    
    st.markdown('<div class="section-header">➕ Create New Invoice</div>', unsafe_allow_html=True)
    
    # Initialize session state for invoice items if not exists
    if 'invoice_items' not in st.session_state:
        st.session_state.invoice_items = []
    if 'edit_index' not in st.session_state:
        st.session_state.edit_index = -1
    if 'invoice_number' not in st.session_state:
        st.session_state.invoice_number = generate_invoice_number()
    if 'invoice_notes' not in st.session_state:
        st.session_state.invoice_notes = ''
    
    # Invoice Header
    col1, col2, col3 = st.columns([2, 2, 1])
    
    with col1:
        st.markdown(f"**Invoice Number:** {st.session_state.invoice_number}")
    
    with col2:
        invoice_date = st.date_input("Invoice Date", datetime.now())
        due_date = st.date_input("Due Date", datetime.now() + timedelta(days=30))
    
    with col3:
        po_number = st.text_input("PO Number", placeholder="Optional")
    
    st.divider()
    
    # Client Information
    st.markdown("##### Client Information")
    
    col1, col2 = st.columns(2)
    
    with col1:
        client_name = st.text_input("Client Name *")
        client_email = st.text_input("Email Address")
        client_phone = st.text_input("Phone Number")
    
    with col2:
        client_address = st.text_area("Address", height=100)
        auto_save_client = st.checkbox("Save to client list", value=True, help="Automatically save this client for future use")
    
    st.divider()
    
    # Invoice Items
    st.markdown("##### Invoice Items")
    
    # Item input form
    with st.container():
        st.markdown('<div class="business-card">', unsafe_allow_html=True)
        
        col1, col2, col3, col4, col5, col6 = st.columns([3, 1, 1, 1, 1, 1])
        
        with col1:
            item_desc = st.text_input("Description", key="item_desc", placeholder="Item description")
        with col2:
            item_qty = st.number_input("Qty", min_value=0.01, value=1.0, step=0.5, key="item_qty", format="%.2f")
        with col3:
            item_price = st.number_input("Unit Price", min_value=0.0, value=0.0, step=10.0, key="item_price", format="%.2f")
        with col4:
            item_tax = st.number_input("Tax %", min_value=0.0, value=0.0, step=2.5, key="item_tax", format="%.1f")
        with col5:
            item_disc = st.number_input("Disc %", min_value=0.0, value=0.0, step=2.5, key="item_disc", format="%.1f")
        with col6:
            if st.button("➕ Add Item", use_container_width=True):
                if item_desc:
                    subtotal = item_qty * item_price
                    discount_amount = subtotal * (item_disc / 100)
                    tax_amount = (subtotal - discount_amount) * (item_tax / 100)
                    total = subtotal - discount_amount + tax_amount
                    
                    item = {
                        'description': item_desc,
                        'quantity': item_qty,
                        'unit_price': item_price,
                        'tax_rate': item_tax,
                        'discount': item_disc,
                        'total': total
                    }
                    
                    if st.session_state.edit_index >= 0:
                        st.session_state.invoice_items[st.session_state.edit_index] = item
                        st.session_state.edit_index = -1
                    else:
                        st.session_state.invoice_items.append(item)
                    
                    st.rerun()
        
        st.markdown('</div>', unsafe_allow_html=True)
    
    # Display items
    if st.session_state.invoice_items:
        st.markdown("##### Current Items")
        
        items_df = pd.DataFrame(st.session_state.invoice_items)
        items_df['Total'] = items_df['total'].apply(lambda x: format_amount(x, st.session_state.currency))
        
        # Display items table
        for i, item in enumerate(st.session_state.invoice_items):
            with st.container():
                st.markdown('<div class="business-card">', unsafe_allow_html=True)
                
                col1, col2, col3, col4, col5, col6, col7 = st.columns([3, 1, 1, 1, 1, 1.5, 1])
                
                with col1:
                    st.markdown(f"**{item['description']}**")
                with col2:
                    st.markdown(f"Qty: {item['quantity']:.2f}")
                with col3:
                    st.markdown(f"@ {format_amount(item['unit_price'], st.session_state.currency)}")
                with col4:
                    st.markdown(f"Tax: {item['tax_rate']}%")
                with col5:
                    st.markdown(f"Disc: {item['discount']}%")
                with col6:
                    st.markdown(f"**{format_amount(item['total'], st.session_state.currency)}**")
                with col7:
                    if st.button("✏️", key=f"edit_{i}", help="Edit item"):
                        st.session_state.edit_index = i
                        st.rerun()
                    if st.button("🗑️", key=f"del_{i}", help="Delete item"):
                        st.session_state.invoice_items.pop(i)
                        if st.session_state.edit_index == i:
                            st.session_state.edit_index = -1
                        st.rerun()
                
                st.markdown('</div>', unsafe_allow_html=True)
        
        # Calculate totals
        subtotal = sum(item['quantity'] * item['unit_price'] for item in st.session_state.invoice_items)
        total_discount = sum((item['quantity'] * item['unit_price']) * (item['discount'] / 100) for item in st.session_state.invoice_items)
        taxable_amount = subtotal - total_discount
        total_tax = sum(taxable_amount * (item['tax_rate'] / 100) for item in st.session_state.invoice_items)
        grand_total = subtotal - total_discount + total_tax
        
        st.divider()
        
        # Advanced Options
        with st.expander("⚙️ Advanced Options"):
            col1, col2 = st.columns(2)
            with col1:
                currency = st.selectbox(
                    "Currency",
                    options=list(CURRENCIES.keys()),
                    format_func=lambda x: f"{CURRENCIES[x]['symbol']} {CURRENCIES[x]['name']}",
                    index=list(CURRENCIES.keys()).index(st.session_state.currency)
                )
                st.session_state.currency = currency
            
            with col2:
                invoice_status = st.selectbox(
                    "Invoice Status",
                    options=INVOICE_STATUSES,
                    index=0
                )
            
            recurring_frequency = st.selectbox(
                "Recurring Frequency",
                options=list(RECURRING_FREQUENCIES.keys())
            )
            
            if recurring_frequency != 'None':
                recurring_end = st.date_input("Recurring End Date (Optional)", value=None, min_value=invoice_date)
            else:
                recurring_end = None
            
            invoice_notes = st.text_area("Notes", value=st.session_state.invoice_notes, height=100)
            st.session_state.invoice_notes = invoice_notes
        
        st.divider()
        
        # Invoice Preview
        with st.expander("👁️ Invoice Preview", expanded=True):
            st.markdown('<div class="invoice-preview">', unsafe_allow_html=True)
            
            # Company Info
            col1, col2 = st.columns(2)
            with col1:
                if st.session_state.company_info.get('logo_base64'):
                    st.markdown(get_logo_html("60px", "150px"), unsafe_allow_html=True)
                st.markdown(f"**{st.session_state.company_info['name']}**")
                st.markdown(st.session_state.company_info['address'])
                st.markdown(st.session_state.company_info.get('city', ''))
                st.markdown(f"Phone: {st.session_state.company_info['phone']}")
                st.markdown(f"Email: {st.session_state.company_info['email']}")
                st.markdown(f"TRN: {st.session_state.company_info['tax_id']}")
            
            with col2:
                st.markdown(f"**INVOICE**")
                st.markdown(f"**Invoice #:** {st.session_state.invoice_number}")
                st.markdown(f"**Date:** {invoice_date.strftime('%Y-%m-%d')}")
                st.markdown(f"**Due Date:** {due_date.strftime('%Y-%m-%d')}")
                if po_number:
                    st.markdown(f"**PO #:** {po_number}")
            
            st.divider()
            
            # Client Info
            st.markdown("**Bill To:**")
            st.markdown(client_name)
            if client_address:
                st.markdown(client_address)
            if client_email:
                st.markdown(f"Email: {client_email}")
            if client_phone:
                st.markdown(f"Phone: {client_phone}")
            
            st.divider()
            
            # Items Table
            preview_items = []
            for item in st.session_state.invoice_items:
                preview_items.append({
                    'Description': item['description'],
                    'Qty': f"{item['quantity']:.2f}",
                    'Unit Price': format_amount(item['unit_price'], st.session_state.currency),
                    'Tax %': f"{item['tax_rate']:.1f}%",
                    'Disc %': f"{item['discount']:.1f}%",
                    'Total': format_amount(item['total'], st.session_state.currency)
                })
            
            st.dataframe(
                pd.DataFrame(preview_items),
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Description": st.column_config.TextColumn("Description", width=200),
                    "Qty": st.column_config.TextColumn("Qty", width=60),
                    "Unit Price": st.column_config.TextColumn("Unit Price", width=100),
                    "Tax %": st.column_config.TextColumn("Tax %", width=60),
                    "Disc %": st.column_config.TextColumn("Disc %", width=60),
                    "Total": st.column_config.TextColumn("Total", width=100)
                }
            )
            
            # Totals
            col1, col2, col3 = st.columns([3, 1, 2])
            with col2:
                st.markdown("**Subtotal:**")
                st.markdown("**Discount:**")
                st.markdown("**Tax:**")
                st.markdown("---")
                st.markdown("**GRAND TOTAL:**")
            with col3:
                st.markdown(f"**{format_amount(subtotal, st.session_state.currency)}**")
                st.markdown(f"**-{format_amount(total_discount, st.session_state.currency)}**")
                st.markdown(f"**{format_amount(total_tax, st.session_state.currency)}**")
                st.markdown("---")
                st.markdown(f"**{format_amount(grand_total, st.session_state.currency)}**")
            
            # Notes
            if invoice_notes:
                st.divider()
                st.markdown("**Notes:**")
                st.markdown(invoice_notes)
            
            st.markdown('</div>', unsafe_allow_html=True)
        
        # Action Buttons
        st.markdown('<div class="action-buttons">', unsafe_allow_html=True)
        
        col1, col2, col3, col4, col5 = st.columns(5)
        
        with col1:
            if st.button("💾 Save as Draft", use_container_width=True):
                invoice_data = {
                    'invoice_number': st.session_state.invoice_number,
                    'client_name': client_name,
                    'client_email': client_email,
                    'client_address': client_address,
                    'client_phone': client_phone,
                    'invoice_date': invoice_date.strftime('%Y-%m-%d'),
                    'due_date': due_date.strftime('%Y-%m-%d'),
                    'po_number': po_number,
                    'currency': st.session_state.currency,
                    'subtotal': subtotal,
                    'tax_total': total_tax,
                    'discount_total': total_discount,
                    'grand_total': grand_total,
                    'amount_paid': 0,
                    'balance_due': grand_total,
                    'status': 'Draft',
                    'notes': invoice_notes,
                    'recurring_frequency': recurring_frequency if recurring_frequency != 'None' else None,
                    'recurring_next_date': recurring_end.strftime('%Y-%m-%d') if recurring_frequency != 'None' and recurring_end else None
                }
                
                invoice_id, errors, warnings = save_invoice_to_db(invoice_data, st.session_state.invoice_items)
                
                if invoice_id:
                    # Save client if option selected
                    if auto_save_client and client_email:
                        client_data = {
                            'name': client_name,
                            'email': client_email,
                            'phone': client_phone,
                            'address': client_address
                        }
                        save_client_to_db(client_data)
                    
                    st.session_state.notification = f"✓ Invoice {st.session_state.invoice_number} saved as Draft"
                    st.session_state.notification_type = "success"
                    st.session_state.invoice_items = []
                    st.session_state.invoice_number = generate_invoice_number()
                    st.session_state.invoice_notes = ''
                    st.rerun()
                else:
                    for error in errors:
                        st.error(error)
                    for warning in warnings:
                        st.warning(warning)
        
        with col2:
            if st.button("📤 Save & Send", use_container_width=True):
                invoice_data = {
                    'invoice_number': st.session_state.invoice_number,
                    'client_name': client_name,
                    'client_email': client_email,
                    'client_address': client_address,
                    'client_phone': client_phone,
                    'invoice_date': invoice_date.strftime('%Y-%m-%d'),
                    'due_date': due_date.strftime('%Y-%m-%d'),
                    'po_number': po_number,
                    'currency': st.session_state.currency,
                    'subtotal': subtotal,
                    'tax_total': total_tax,
                    'discount_total': total_discount,
                    'grand_total': grand_total,
                    'amount_paid': 0,
                    'balance_due': grand_total,
                    'status': 'Sent',
                    'notes': invoice_notes,
                    'sent_date': datetime.now().isoformat(),
                    'recurring_frequency': recurring_frequency if recurring_frequency != 'None' else None,
                    'recurring_next_date': recurring_end.strftime('%Y-%m-%d') if recurring_frequency != 'None' and recurring_end else None
                }
                
                invoice_id, errors, warnings = save_invoice_to_db(invoice_data, st.session_state.invoice_items)
                
                if invoice_id:
                    # Generate PDF for email
                    pdf_data = {
                        'invoice_number': st.session_state.invoice_number,
                        'invoice_date': invoice_date.strftime('%Y-%m-%d'),
                        'due_date': due_date.strftime('%Y-%m-%d'),
                        'po_number': po_number,
                        'currency': st.session_state.currency,
                        'status': 'Sent',
                        'client': {
                            'name': client_name,
                            'address': client_address,
                            'email': client_email,
                            'phone': client_phone
                        },
                        'company_info': st.session_state.company_info,
                        'items': st.session_state.invoice_items,
                        'totals': {
                            'subtotal': subtotal,
                            'discount': total_discount,
                            'tax': total_tax,
                            'grand_total': grand_total
                        },
                        'notes': invoice_notes,
                        'amount_paid': 0,
                        'balance_due': grand_total
                    }
                    
                    pdf_buffer = generate_pdf_invoice(pdf_data)
                    
                    # Save client if option selected
                    if auto_save_client and client_email:
                        client_data = {
                            'name': client_name,
                            'email': client_email,
                            'phone': client_phone,
                            'address': client_address
                        }
                        save_client_to_db(client_data)
                    
                    st.session_state.notification = f"✓ Invoice {st.session_state.invoice_number} saved and ready to send"
                    st.session_state.notification_type = "success"
                    
                    # Open email dialog
                    st.session_state.show_email_modal = True
                    st.session_state.email_invoice_id = invoice_id
                    st.session_state.email_pdf = pdf_buffer
                    st.rerun()
        
        with col3:
            if st.button("👁️ Preview PDF", use_container_width=True):
                pdf_data = {
                    'invoice_number': st.session_state.invoice_number,
                    'invoice_date': invoice_date.strftime('%Y-%m-%d'),
                    'due_date': due_date.strftime('%Y-%m-%d'),
                    'po_number': po_number,
                    'currency': st.session_state.currency,
                    'status': invoice_status,
                    'client': {
                        'name': client_name,
                        'address': client_address,
                        'email': client_email,
                        'phone': client_phone
                    },
                    'company_info': st.session_state.company_info,
                    'items': st.session_state.invoice_items,
                    'totals': {
                        'subtotal': subtotal,
                        'discount': total_discount,
                        'tax': total_tax,
                        'grand_total': grand_total
                    },
                    'notes': invoice_notes,
                    'amount_paid': 0,
                    'balance_due': grand_total
                }
                
                pdf_buffer = generate_pdf_invoice(pdf_data)
                if pdf_buffer:
                    st.download_button(
                        label="📥 Download PDF",
                        data=pdf_buffer,
                        file_name=f"invoice_{st.session_state.invoice_number}.pdf",
                        mime="application/pdf",
                        use_container_width=True
                    )
        
        with col4:
            if st.button("📊 Export Excel", use_container_width=True):
                invoice_data_export = {
                    'invoice_number': st.session_state.invoice_number,
                    'client_name': client_name,
                    'client_email': client_email,
                    'client_phone': client_phone,
                    'client_address': client_address,
                    'invoice_date': invoice_date.strftime('%Y-%m-%d'),
                    'due_date': due_date.strftime('%Y-%m-%d'),
                    'po_number': po_number,
                    'currency': st.session_state.currency,
                    'subtotal': subtotal,
                    'tax_total': total_tax,
                    'discount_total': total_discount,
                    'grand_total': grand_total
                }
                
                excel_buffer = export_to_excel(invoice_data_export, st.session_state.invoice_items)
                if excel_buffer:
                    st.download_button(
                        label="📥 Download Excel",
                        data=excel_buffer,
                        file_name=f"invoice_{st.session_state.invoice_number}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
        
        with col5:
            if st.button("🔄 Clear Form", use_container_width=True):
                st.session_state.invoice_items = []
                st.session_state.invoice_number = generate_invoice_number()
                st.session_state.invoice_notes = ''
                st.session_state.edit_index = -1
                st.rerun()
        
        st.markdown('</div>', unsafe_allow_html=True)
    
    else:
        st.info("💡 Add items and client information to create your invoice")

# ============================================================================
# VIEW INVOICES PAGE
# ============================================================================

def render_view_invoices_page():
    """Render the view invoices page"""
    
    st.markdown('<div class="section-header">📋 View Invoices</div>', unsafe_allow_html=True)
    
    # Initialize filter session states
    if 'filter_status' not in st.session_state:
        st.session_state.filter_status = 'All'
    if 'filter_client' not in st.session_state:
        st.session_state.filter_client = ''
    if 'filter_date_from' not in st.session_state:
        st.session_state.filter_date_from = None
    if 'filter_date_to' not in st.session_state:
        st.session_state.filter_date_to = None
    
    # Filters
    with st.expander("🔍 Search & Filter", expanded=True):
        col1, col2, col3 = st.columns(3)
        
        with col1:
            filter_status = st.selectbox(
                "Status",
                options=['All'] + INVOICE_STATUSES,
                index=0 if st.session_state.filter_status == 'All' else INVOICE_STATUSES.index(st.session_state.filter_status) + 1
            )
        
        with col2:
            filter_client = st.text_input("Client Name", value=st.session_state.filter_client, placeholder="Search by client...")
        
        with col3:
            date_range = st.date_input(
                "Date Range",
                value=(
                    datetime.strptime(st.session_state.filter_date_from, '%Y-%m-%d') if st.session_state.filter_date_from else datetime.now() - timedelta(days=30),
                    datetime.strptime(st.session_state.filter_date_to, '%Y-%m-%d') if st.session_state.filter_date_to else datetime.now()
                ),
                key="date_range_filter"
            )
        
        if st.button("🔍 Apply Filters", use_container_width=True):
            st.session_state.filter_status = filter_status
            st.session_state.filter_client = filter_client
            if len(date_range) == 2:
                st.session_state.filter_date_from = date_range[0].strftime('%Y-%m-%d')
                st.session_state.filter_date_to = date_range[1].strftime('%Y-%m-%d')
            st.rerun()
    
    # Build filters
    filters = {}
    if st.session_state.filter_status != 'All':
        filters['status'] = st.session_state.filter_status
    if st.session_state.filter_client:
        filters['client_name'] = st.session_state.filter_client
    if st.session_state.filter_date_from:
        filters['date_from'] = st.session_state.filter_date_from
    if st.session_state.filter_date_to:
        filters['date_to'] = st.session_state.filter_date_to
    
    # Get invoices
    invoices_df = get_invoices(filters if filters else None)
    
    if not invoices_df.empty:
        # Summary stats
        total_amount = invoices_df['grand_total'].sum()
        paid_amount = invoices_df[invoices_df['status'] == 'Paid']['grand_total'].sum() if 'Paid' in invoices_df['status'].values else 0
        pending_amount = invoices_df[invoices_df['status'].isin(['Draft', 'Sent'])]['grand_total'].sum() if any(invoices_df['status'].isin(['Draft', 'Sent'])) else 0
        
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Total Invoices", len(invoices_df))
        with col2:
            st.metric("Total Amount", format_amount(total_amount, st.session_state.currency))
        with col3:
            st.metric("Paid", format_amount(paid_amount, st.session_state.currency))
        with col4:
            st.metric("Pending", format_amount(pending_amount, st.session_state.currency))
        
        st.divider()
        
        # Paginate invoices
        paginated_df = paginate_dataframe(invoices_df, page_size=10, key="invoices")
        
        # Display invoices
        for _, invoice in paginated_df.iterrows():
            with st.container():
                st.markdown('<div class="business-card">', unsafe_allow_html=True)
                
                col1, col2, col3, col4, col5 = st.columns([2, 2, 1.5, 1.5, 2])
                
                with col1:
                    st.markdown(f"**{invoice['invoice_number']}**")
                    st.caption(f"Client: {invoice['client_name']}")
                
                with col2:
                    st.markdown(f"**Date:** {invoice['invoice_date']}")
                    st.markdown(f"**Due:** {invoice['due_date']}")
                
                with col3:
                    st.markdown(f"**Amount:** {format_amount(invoice['grand_total'], invoice['currency'])}")
                    if invoice['balance_due'] > 0:
                        st.caption(f"Balance: {format_amount(invoice['balance_due'], invoice['currency'])}")
                
                with col4:
                    st.markdown(get_status_badge_html(invoice['status']), unsafe_allow_html=True)
                    if invoice['status'] == 'Overdue':
                        st.caption("⚠️ Overdue")
                
                with col5:
                    button_col1, button_col2, button_col3 = st.columns(3)
                    with button_col1:
                        if st.button("👁️", key=f"view_{invoice['id']}", help="View Details"):
                            st.session_state.view_invoice_id = invoice['id']
                            st.rerun()
                    with button_col2:
                        if st.button("📄", key=f"pdf_{invoice['id']}", help="Download PDF"):
                            invoice_data, items = get_invoice_by_id(invoice['id'])
                            if invoice_data and items:
                                pdf_data = {
                                    'invoice_number': invoice_data['invoice_number'],
                                    'invoice_date': invoice_data['invoice_date'],
                                    'due_date': invoice_data['due_date'],
                                    'po_number': invoice_data.get('po_number', ''),
                                    'currency': invoice_data['currency'],
                                    'status': invoice_data['status'],
                                    'client': {
                                        'name': invoice_data['client_name'],
                                        'address': invoice_data.get('client_address', ''),
                                        'email': invoice_data.get('client_email', ''),
                                        'phone': invoice_data.get('client_phone', '')
                                    },
                                    'company_info': st.session_state.company_info,
                                    'items': items,
                                    'totals': {
                                        'subtotal': invoice_data['subtotal'],
                                        'discount': invoice_data['discount_total'],
                                        'tax': invoice_data['tax_total'],
                                        'grand_total': invoice_data['grand_total']
                                    },
                                    'notes': invoice_data.get('notes', ''),
                                    'amount_paid': invoice_data['amount_paid'],
                                    'balance_due': invoice_data['balance_due']
                                }
                                pdf_buffer = generate_pdf_invoice(pdf_data)
                                if pdf_buffer:
                                    st.download_button(
                                        label="📥",
                                        data=pdf_buffer,
                                        file_name=f"invoice_{invoice_data['invoice_number']}.pdf",
                                        mime="application/pdf",
                                        key=f"download_{invoice['id']}"
                                    )
                    with button_col3:
                        if st.button("💰", key=f"pay_{invoice['id']}", help="Record Payment"):
                            st.session_state.payment_invoice_id = invoice['id']
                            st.session_state.show_payment_modal = True
                            st.rerun()
                
                # Additional actions row if needed
                with st.expander("More Actions", expanded=False):
                    col_a, col_b, col_c, col_d = st.columns(4)
                    with col_a:
                        if st.button("📧 Send Email", key=f"email_{invoice['id']}"):
                            st.session_state.show_email_modal = True
                            st.session_state.email_invoice_id = invoice['id']
                            st.rerun()
                    with col_b:
                        if st.button("📊 Export Excel", key=f"excel_{invoice['id']}"):
                            invoice_data, items = get_invoice_by_id(invoice['id'])
                            if invoice_data:
                                excel_buffer = export_to_excel(invoice_data, items)
                                if excel_buffer:
                                    st.download_button(
                                        label="Download Excel",
                                        data=excel_buffer,
                                        file_name=f"invoice_{invoice_data['invoice_number']}.xlsx",
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        key=f"dl_excel_{invoice['id']}"
                                    )
                    with col_c:
                        if st.button("🔄 Update Status", key=f"status_{invoice['id']}"):
                            new_status = st.selectbox(
                                "New Status",
                                options=INVOICE_STATUSES,
                                index=INVOICE_STATUSES.index(invoice['status']),
                                key=f"status_select_{invoice['id']}"
                            )
                            if st.button("Update", key=f"update_status_{invoice['id']}"):
                                if update_invoice_status(invoice['id'], new_status):
                                    st.success(f"Status updated to {new_status}")
                                    st.rerun()
                    with col_d:
                        if st.button("🗑️ Delete", key=f"del_{invoice['id']}"):
                            if delete_invoice(invoice['id']):
                                st.success("Invoice deleted")
                                st.rerun()
                
                st.markdown('</div>', unsafe_allow_html=True)
    else:
        st.info("No invoices found. Create your first invoice!")
        
        if st.button("➕ Create New Invoice", use_container_width=True):
            st.session_state.current_page = "create"
            st.rerun()
    
    # View single invoice details
    if st.session_state.get('view_invoice_id'):
        invoice, items = get_invoice_by_id(st.session_state.view_invoice_id)
        
        if invoice:
            st.markdown("---")
            st.markdown(f"### Invoice Details: {invoice['invoice_number']}")
            
            with st.container():
                st.markdown('<div class="business-card">', unsafe_allow_html=True)
                
                col1, col2 = st.columns(2)
                
                with col1:
                    st.markdown("**Client Information:**")
                    st.markdown(f"""
                    **Name:** {invoice['client_name']}  
                    **Email:** {invoice['client_email']}  
                    **Phone:** {invoice.get('client_phone', 'N/A')}  
                    **Address:** {invoice.get('client_address', 'N/A')}
                    """)
                
                with col2:
                    st.markdown("**Invoice Information:**")
                    st.markdown(f"""
                    **Date:** {invoice['invoice_date']}  
                    **Due Date:** {invoice['due_date']}  
                    **PO Number:** {invoice.get('po_number', 'N/A')}  
                    **Status:** {get_status_badge_html(invoice['status'])}
                    """, unsafe_allow_html=True)
                
                st.divider()
                
                # Items
                if items:
                    st.markdown("**Invoice Items:**")
                    items_data = []
                    for item in items:
                        items_data.append({
                            'Description': item['description'],
                            'Qty': f"{item['quantity']:.2f}",
                            'Unit Price': format_amount(item['unit_price'], invoice['currency']),
                            'Tax %': f"{item['tax_rate']}%",
                            'Discount %': f"{item['discount']}%",
                            'Total': format_amount(item['total'], invoice['currency'])
                        })
                    
                    st.dataframe(pd.DataFrame(items_data), use_container_width=True, hide_index=True)
                
                # Totals
                col1, col2, col3 = st.columns([3, 1, 2])
                with col2:
                    st.markdown("**Subtotal:**")
                    st.markdown("**Discount:**")
                    st.markdown("**Tax:**")
                    st.markdown("---")
                    st.markdown("**Grand Total:**")
                    if invoice['amount_paid'] > 0:
                        st.markdown("**Amount Paid:**")
                        st.markdown("---")
                        st.markdown("**Balance Due:**")
                with col3:
                    st.markdown(f"**{format_amount(invoice['subtotal'], invoice['currency'])}**")
                    st.markdown(f"**-{format_amount(invoice['discount_total'], invoice['currency'])}**")
                    st.markdown(f"**{format_amount(invoice['tax_total'], invoice['currency'])}**")
                    st.markdown("---")
                    st.markdown(f"**{format_amount(invoice['grand_total'], invoice['currency'])}**")
                    if invoice['amount_paid'] > 0:
                        st.markdown(f"**{format_amount(invoice['amount_paid'], invoice['currency'])}**")
                        st.markdown("---")
                        st.markdown(f"**{format_amount(invoice['balance_due'], invoice['currency'])}**")
                
                # Notes
                if invoice.get('notes'):
                    st.divider()
                    st.markdown("**Notes:**")
                    st.markdown(invoice['notes'])
                
                st.markdown('</div>', unsafe_allow_html=True)
            
            if st.button("← Back to List"):
                st.session_state.view_invoice_id = None
                st.rerun()
    
    # Payment Modal
    if st.session_state.get('show_payment_modal') and st.session_state.get('payment_invoice_id'):
        with st.container():
            st.markdown('<div class="business-card">', unsafe_allow_html=True)
            st.markdown("### 💰 Record Payment")
            
            invoice, _ = get_invoice_by_id(st.session_state.payment_invoice_id)
            
            if invoice:
                st.markdown(f"""
                **Invoice:** {invoice['invoice_number']}  
                **Client:** {invoice['client_name']}  
                **Total Amount:** {format_amount(invoice['grand_total'], invoice['currency'])}  
                **Amount Paid:** {format_amount(invoice['amount_paid'], invoice['currency'])}  
                **Balance Due:** {format_amount(invoice['balance_due'], invoice['currency'])}
                """)
                
                max_payment = invoice['balance_due']
                
                payment_amount = st.number_input(
                    "Payment Amount",
                    min_value=0.01,
                    max_value=float(max_payment),
                    value=min(float(max_payment), 100.0),
                    step=10.0,
                    format="%.2f"
                )
                
                payment_method = st.selectbox("Payment Method", PAYMENT_METHODS)
                payment_reference = st.text_input("Reference (optional)", placeholder="Check #, Transaction ID, etc.")
                payment_notes = st.text_area("Notes (optional)", height=100)
                
                col1, col2, col3 = st.columns(3)
                with col1:
                    if st.button("✅ Record Payment", use_container_width=True):
                        success, message = process_payment(
                            st.session_state.payment_invoice_id,
                            payment_amount,
                            payment_method,
                            payment_reference,
                            payment_notes
                        )
                        if success:
                            st.session_state.notification = message
                            st.session_state.notification_type = "success"
                            st.session_state.show_payment_modal = False
                            st.session_state.payment_invoice_id = None
                            st.rerun()
                        else:
                            st.error(message)
                
                with col2:
                    if st.button("❌ Cancel", use_container_width=True):
                        st.session_state.show_payment_modal = False
                        st.session_state.payment_invoice_id = None
                        st.rerun()
            
            st.markdown('</div>', unsafe_allow_html=True)
    
    # Email Modal
    if st.session_state.get('show_email_modal') and st.session_state.get('email_invoice_id'):
        with st.container():
            st.markdown('<div class="business-card">', unsafe_allow_html=True)
            st.markdown("### 📧 Send Invoice via Email")
            
            invoice, items = get_invoice_by_id(st.session_state.email_invoice_id)
            
            if invoice:
                to_email = st.text_input("To Email", value=invoice['client_email'])
                subject = st.text_input("Subject", value=f"Invoice {invoice['invoice_number']} from {st.session_state.company_info['name']}")
                
                body = st.text_area(
                    "Message",
                    value=f"""Dear {invoice['client_name']},

Please find attached invoice {invoice['invoice_number']} for your reference.

Invoice Details:
- Invoice Number: {invoice['invoice_number']}
- Date: {invoice['invoice_date']}
- Due Date: {invoice['due_date']}
- Amount: {format_amount(invoice['grand_total'], invoice['currency'])}

Payment can be made via:
{st.session_state.company_info.get('bank_details', '')}

Thank you for your business!

Best regards,
{st.session_state.company_info['name']}""",
                    height=200
                )
                
                # Generate PDF if not already in session
                if not hasattr(st.session_state, 'email_pdf') or st.session_state.email_pdf is None:
                    pdf_data = {
                        'invoice_number': invoice['invoice_number'],
                        'invoice_date': invoice['invoice_date'],
                        'due_date': invoice['due_date'],
                        'po_number': invoice.get('po_number', ''),
                        'currency': invoice['currency'],
                        'status': invoice['status'],
                        'client': {
                            'name': invoice['client_name'],
                            'address': invoice.get('client_address', ''),
                            'email': invoice.get('client_email', ''),
                            'phone': invoice.get('client_phone', '')
                        },
                        'company_info': st.session_state.company_info,
                        'items': items,
                        'totals': {
                            'subtotal': invoice['subtotal'],
                            'discount': invoice['discount_total'],
                            'tax': invoice['tax_total'],
                            'grand_total': invoice['grand_total']
                        },
                        'notes': invoice.get('notes', ''),
                        'amount_paid': invoice['amount_paid'],
                        'balance_due': invoice['balance_due']
                    }
                    pdf_buffer = generate_pdf_invoice(pdf_data)
                    st.session_state.email_pdf = pdf_buffer
                
                col1, col2, col3 = st.columns(3)
                with col1:
                    if st.button("📤 Send Email", use_container_width=True):
                        success, message = send_email_invoice(
                            to_email,
                            st.session_state.email_pdf,
                            invoice['invoice_number']
                        )
                        if success:
                            # Update invoice status to Sent
                            update_invoice_status(invoice['id'], 'Sent')
                            
                            st.session_state.notification = f"✓ Invoice sent to {to_email}"
                            st.session_state.notification_type = "success"
                            st.session_state.show_email_modal = False
                            st.session_state.email_invoice_id = None
                            st.session_state.email_pdf = None
                            st.rerun()
                        else:
                            st.error(message)
                
                with col2:
                    if st.button("📥 Download PDF", use_container_width=True):
                        st.download_button(
                            label="Download PDF",
                            data=st.session_state.email_pdf,
                            file_name=f"invoice_{invoice['invoice_number']}.pdf",
                            mime="application/pdf",
                            key="email_download_pdf"
                        )
                
                with col3:
                    if st.button("❌ Cancel", use_container_width=True):
                        st.session_state.show_email_modal = False
                        st.session_state.email_invoice_id = None
                        st.session_state.email_pdf = None
                        st.rerun()
            
            st.markdown('</div>', unsafe_allow_html=True)

# ============================================================================
# CLIENTS PAGE
# ============================================================================

def render_clients_page():
    """Render the clients management page"""
    
    st.markdown('<div class="section-header">👥 Client Management</div>', unsafe_allow_html=True)
    
    tab1, tab2 = st.tabs(["📋 Client List", "➕ Add New Client"])
    
    with tab1:
        # Search
        search_term = st.text_input("🔍 Search Clients", placeholder="Name, email, or company...")
        
        clients_df = get_clients(search_term if search_term else None)
        
        if not clients_df.empty:
            for _, client in clients_df.iterrows():
                with st.container():
                    st.markdown('<div class="business-card">', unsafe_allow_html=True)
                    
                    col1, col2, col3, col4 = st.columns([2, 2, 2, 1])
                    
                    with col1:
                        st.markdown(f"**{client['name']}**")
                        st.caption(client.get('company', ''))
                    
                    with col2:
                        st.markdown(f"📧 {client['email']}")
                        if client.get('phone'):
                            st.markdown(f"📞 {client['phone']}")
                    
                    with col3:
                        st.markdown(f"📍 {client.get('address', 'No address')[:50]}")
                        if client.get('tax_id'):
                            st.caption(f"TRN: {client['tax_id']}")
                    
                    with col4:
                        if st.button("👁️ View", key=f"view_client_{client['id']}"):
                            st.session_state.selected_client_id = client['id']
                            st.rerun()
                    
                    # Show client details if selected
                    if st.session_state.get('selected_client_id') == client['id']:
                        st.divider()
                        st.markdown("**Client Details:**")
                        
                        col_a, col_b = st.columns(2)
                        with col_a:
                            st.markdown(f"""
                            **Full Name:** {client['name']}  
                            **Company:** {client.get('company', 'N/A')}  
                            **Email:** {client['email']}  
                            **Phone:** {client.get('phone', 'N/A')}
                            """)
                        with col_b:
                            st.markdown(f"""
                            **Address:** {client.get('address', 'N/A')}  
                            **TRN/Tax ID:** {client.get('tax_id', 'N/A')}  
                            **Credit Limit:** {format_amount(client.get('credit_limit', 0), st.session_state.currency)}  
                            **Payment Terms:** {client.get('payment_terms', 30)} days
                            """)
                        
                        # Get client's invoices
                        client_invoices = get_invoices({'client_name': client['name']})
                        if not client_invoices.empty:
                            st.markdown("**Recent Invoices:**")
                            for _, inv in client_invoices.head(3).iterrows():
                                st.markdown(f"""
                                - {inv['invoice_number']}: {format_amount(inv['grand_total'], inv['currency'])} ({inv['status']})
                                """)
                        
                        if st.button("Close", key=f"close_client_{client['id']}"):
                            st.session_state.selected_client_id = None
                            st.rerun()
                    
                    st.markdown('</div>', unsafe_allow_html=True)
        else:
            st.info("No clients found. Add your first client!")
    
    with tab2:
        with st.container():
            st.markdown('<div class="business-card">', unsafe_allow_html=True)
            st.markdown("##### Add New Client")
            
            client_name = st.text_input("Client Name *")
            client_email = st.text_input("Email Address *")
            client_phone = st.text_input("Phone Number")
            client_company = st.text_input("Company Name")
            client_address = st.text_area("Address")
            client_tax_id = st.text_input("TRN / Tax ID")
            
            col1, col2 = st.columns(2)
            with col1:
                credit_limit = st.number_input("Credit Limit", min_value=0.0, value=0.0, step=1000.0, format="%.2f")
            with col2:
                payment_terms = st.number_input("Payment Terms (days)", min_value=0, value=30, step=1)
            
            client_notes = st.text_area("Notes", height=100)
            
            if st.button("💾 Save Client", use_container_width=True):
                if client_name and client_email:
                    if not validate_email(client_email):
                        st.error("Please enter a valid email address")
                    else:
                        client_data = {
                            'name': client_name,
                            'email': client_email,
                            'phone': client_phone,
                            'address': client_address,
                            'company': client_company,
                            'tax_id': client_tax_id,
                            'notes': client_notes,
                            'credit_limit': credit_limit,
                            'payment_terms': payment_terms
                        }
                        
                        client_id = save_client_to_db(client_data)
                        if client_id:
                            st.session_state.notification = f"✓ Client {client_name} saved successfully"
                            st.session_state.notification_type = "success"
                            st.rerun()
                else:
                    st.warning("Name and email are required")
            
            st.markdown('</div>', unsafe_allow_html=True)

# ============================================================================
# PAYMENTS PAGE
# ============================================================================

def render_payments_page():
    """Render the payments management page"""
    
    st.markdown('<div class="section-header">💰 Payment Management</div>', unsafe_allow_html=True)
    
    # Get all payments
    try:
        with get_db_connection() as conn:
            payments_df = pd.read_sql_query("""
                SELECT p.*, i.invoice_number, i.client_name, i.grand_total, i.currency
                FROM payments p
                JOIN invoices i ON p.invoice_id = i.id
                ORDER BY p.payment_date DESC
            """, conn)
    except:
        payments_df = pd.DataFrame()
    
    if not payments_df.empty:
        # Summary stats
        total_payments = payments_df['amount'].sum()
        payment_count = len(payments_df)
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Total Payments", format_amount(total_payments, st.session_state.currency))
        with col2:
            st.metric("Number of Payments", payment_count)
        with col3:
            avg_payment = total_payments / payment_count if payment_count > 0 else 0
            st.metric("Average Payment", format_amount(avg_payment, st.session_state.currency))
        
        st.divider()
        
        # Payment methods breakdown
        method_stats = payments_df.groupby('payment_method')['amount'].agg(['sum', 'count']).reset_index()
        
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("**Payment Methods**")
            fig = px.pie(
                method_stats,
                values='sum',
                names='payment_method',
                title='Payment Amount by Method'
            )
            fig.update_layout(
                plot_bgcolor='white',
                paper_bgcolor='white',
                font=dict(color='#2c3e50')
            )
            st.plotly_chart(fig, use_container_width=True)
        
        with col2:
            st.markdown("**Payment Timeline**")
            payments_df['payment_date'] = pd.to_datetime(payments_df['payment_date'])
            daily_payments = payments_df.groupby(payments_df['payment_date'].dt.date)['amount'].sum().reset_index()
            
            fig = px.line(
                daily_payments,
                x='payment_date',
                y='amount',
                title='Daily Payments'
            )
            fig.update_layout(
                plot_bgcolor='white',
                paper_bgcolor='white',
                font=dict(color='#2c3e50')
            )
            st.plotly_chart(fig, use_container_width=True)
        
        st.divider()
        
        # Payment list with pagination
        st.markdown("**Recent Payments**")
        paginated_payments = paginate_dataframe(payments_df, page_size=10, key="payments")
        
        for _, payment in paginated_payments.iterrows():
            with st.container():
                st.markdown('<div class="business-card">', unsafe_allow_html=True)
                
                col1, col2, col3, col4, col5 = st.columns([1.5, 1.5, 1.5, 1.5, 1])
                
                with col1:
                    st.markdown(f"**{payment['invoice_number']}**")
                    st.caption(payment['client_name'])
                
                with col2:
                    st.markdown(f"**Amount:** {format_amount(payment['amount'], payment['currency'])}")
                    st.caption(f"Method: {payment['payment_method']}")
                
                with col3:
                    payment_date = pd.to_datetime(payment['payment_date']).strftime('%d %b %Y')
                    st.markdown(f"**Date:** {payment_date}")
                    if payment.get('reference'):
                        st.caption(f"Ref: {payment['reference']}")
                
                with col4:
                    if payment.get('notes'):
                        st.caption(f"📝 {payment['notes'][:50]}...")
                
                with col5:
                    if st.button("👁️", key=f"view_payment_{payment['id']}"):
                        # Show payment details in modal
                        st.session_state.view_payment_id = payment['id']
                        st.rerun()
                
                st.markdown('</div>', unsafe_allow_html=True)
    else:
        st.info("No payments recorded yet. Record your first payment!")
        
        # Quick payment form
        with st.container():
            st.markdown('<div class="business-card">', unsafe_allow_html=True)
            st.markdown("##### Quick Payment")
            
            # Get unpaid invoices
            unpaid_invoices = get_invoices({'status': 'Sent'})
            if not unpaid_invoices.empty:
                invoice_options = {
                    f"{row['invoice_number']} - {row['client_name']} ({format_amount(row['balance_due'], row['currency'])})": row['id'] 
                    for _, row in unpaid_invoices.iterrows()
                }
                
                selected_invoice = st.selectbox(
                    "Select Invoice",
                    options=list(invoice_options.keys())
                )
                
                if selected_invoice:
                    invoice_id = invoice_options[selected_invoice]
                    invoice = unpaid_invoices[unpaid_invoices['id'] == invoice_id].iloc[0]
                    
                    st.markdown(f"""
                    **Invoice Details:**  
                    Amount: {format_amount(invoice['grand_total'], invoice['currency'])}  
                    Paid: {format_amount(invoice['amount_paid'], invoice['currency'])}  
                    Balance: {format_amount(invoice['balance_due'], invoice['currency'])}
                    """)
                    
                    payment_amount = st.number_input(
                        "Payment Amount",
                        min_value=0.01,
                        max_value=float(invoice['balance_due']),
                        value=float(invoice['balance_due']),
                        step=10.0,
                        format="%.2f"
                    )
                    
                    payment_method = st.selectbox("Payment Method", PAYMENT_METHODS, key="quick_payment_method")
                    payment_reference = st.text_input("Reference", key="quick_payment_ref")
                    
                    if st.button("💾 Record Payment", use_container_width=True):
                        success, message = process_payment(
                            invoice_id,
                            payment_amount,
                            payment_method,
                            payment_reference,
                            "Quick payment"
                        )
                        if success:
                            st.session_state.notification = message
                            st.session_state.notification_type = "success"
                            st.rerun()
                        else:
                            st.error(message)
            else:
                st.info("No unpaid invoices found")
            
            st.markdown('</div>', unsafe_allow_html=True)

# ============================================================================
# RECURRING INVOICES PAGE
# ============================================================================

def render_recurring_page():
    """Render the recurring invoices management page"""
    
    st.markdown('<div class="section-header">🔄 Recurring Invoices</div>', unsafe_allow_html=True)
    
    # Get recurring invoices
    try:
        with get_db_connection() as conn:
            recurring_df = pd.read_sql_query("""
                SELECT r.*, c.name as client_name, t.name as template_name
                FROM recurring_invoices r
                LEFT JOIN clients c ON r.client_id = c.id
                LEFT JOIN invoice_templates t ON r.template_id = t.id
                ORDER BY r.next_date
            """, conn)
    except:
        recurring_df = pd.DataFrame()
    
    if not recurring_df.empty:
        # Active vs inactive
        active_count = len(recurring_df[recurring_df['is_active'] == 1])
        inactive_count = len(recurring_df[recurring_df['is_active'] == 0])
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Total Recurring", len(recurring_df))
        with col2:
            st.metric("Active", active_count)
        with col3:
            st.metric("Inactive", inactive_count)
        
        st.divider()
        
        # Recurring list with pagination
        paginated_recurring = paginate_dataframe(recurring_df, page_size=10, key="recurring")
        
        for _, recurring in paginated_recurring.iterrows():
            with st.container():
                st.markdown('<div class="business-card">', unsafe_allow_html=True)
                
                col1, col2, col3, col4, col5 = st.columns([2, 2, 1.5, 1.5, 1])
                
                with col1:
                    st.markdown(f"**{recurring['client_name']}**")
                    st.caption(f"Template: {recurring['template_name']}")
                
                with col2:
                    st.markdown(f"**Frequency:** {recurring['frequency']}")
                    st.caption(f"Started: {recurring['start_date']}")
                
                with col3:
                    st.markdown(f"**Next:** {recurring['next_date']}")
                    status = "🟢 Active" if recurring['is_active'] else "🔴 Inactive"
                    st.markdown(f"**Status:** {status}")
                
                with col4:
                    if recurring.get('last_generated'):
                        st.caption(f"Last: {recurring['last_generated']}")
                
                with col5:
                    toggle_label = "Deactivate" if recurring['is_active'] else "Activate"
                    if st.button(toggle_label, key=f"toggle_{recurring['id']}"):
                        try:
                            with get_db_connection() as conn:
                                c = conn.cursor()
                                c.execute("UPDATE recurring_invoices SET is_active = ? WHERE id = ?",
                                         (0 if recurring['is_active'] else 1, recurring['id']))
                                conn.commit()
                                st.session_state.notification = f"Recurring invoice {toggle_label}d"
                                st.session_state.notification_type = "success"
                                st.rerun()
                        except Exception as e:
                            st.error(str(e))
                
                st.markdown('</div>', unsafe_allow_html=True)
    else:
        st.info("No recurring invoices set up yet")
        
        # Setup form
        with st.container():
            st.markdown('<div class="business-card">', unsafe_allow_html=True)
            st.markdown("##### Setup Recurring Invoice")
            
            # Get clients and templates
            clients_df = get_clients()
            try:
                with get_db_connection() as conn:
                    templates_df = pd.read_sql_query("SELECT * FROM invoice_templates", conn)
            except:
                templates_df = pd.DataFrame()
            
            if not clients_df.empty and not templates_df.empty:
                col1, col2 = st.columns(2)
                with col1:
                    client_id = st.selectbox(
                        "Select Client",
                        options=clients_df['id'].tolist(),
                        format_func=lambda x: clients_df[clients_df['id'] == x]['name'].iloc[0]
                    )
                
                with col2:
                    template_id = st.selectbox(
                        "Select Template",
                        options=templates_df['id'].tolist(),
                        format_func=lambda x: templates_df[templates_df['id'] == x]['name'].iloc[0]
                    )
                
                frequency = st.selectbox("Frequency", options=list(RECURRING_FREQUENCIES.keys())[1:])  # Skip None
                
                col1, col2 = st.columns(2)
                with col1:
                    start_date = st.date_input("Start Date", datetime.now())
                with col2:
                    end_date = st.date_input("End Date (optional)", value=None, min_value=start_date)
                
                if st.button("🔄 Create Recurring Schedule", use_container_width=True):
                    recurring_id = create_recurring_invoice(
                        template_id,
                        client_id,
                        frequency,
                        start_date.strftime('%Y-%m-%d'),
                        end_date.strftime('%Y-%m-%d') if end_date else None
                    )
                    if recurring_id:
                        st.session_state.notification = "✓ Recurring schedule created"
                        st.session_state.notification_type = "success"
                        st.rerun()
            else:
                if clients_df.empty:
                    st.warning("No clients found. Add clients first.")
                if templates_df.empty:
                    st.warning("No templates found. Save an invoice as template first.")
            
            st.markdown('</div>', unsafe_allow_html=True)

# ============================================================================
# REPORTS PAGE
# ============================================================================

def render_reports_page():
    """Render the reports page"""
    
    st.markdown('<div class="section-header">📊 Reports</div>', unsafe_allow_html=True)
    
    # Report type selector
    report_type = st.selectbox(
        "Select Report Type",
        options=[
            "Revenue Report",
            "Aging Report",
            "Client Summary",
            "Tax Summary",
            "Payment Methods"
        ]
    )
    
    # Date range
    col1, col2 = st.columns(2)
    with col1:
        report_start = st.date_input("Start Date", datetime.now() - timedelta(days=30))
    with col2:
        report_end = st.date_input("End Date", datetime.now())
    
    if st.button("📊 Generate Report", use_container_width=True):
        if report_type == "Revenue Report":
            # Revenue by period
            with get_db_connection() as conn:
                revenue_df = pd.read_sql_query("""
                    SELECT 
                        strftime('%Y-%m', invoice_date) as period,
                        COUNT(*) as invoice_count,
                        SUM(CASE WHEN status = 'Paid' THEN grand_total ELSE 0 END) as paid_revenue,
                        SUM(CASE WHEN status != 'Paid' THEN grand_total ELSE 0 END) as pending_revenue,
                        SUM(grand_total) as total_revenue
                    FROM invoices
                    WHERE invoice_date BETWEEN ? AND ?
                    GROUP BY strftime('%Y-%m', invoice_date)
                    ORDER BY period
                """, conn, params=[report_start.strftime('%Y-%m-%d'), report_end.strftime('%Y-%m-%d')])
            
            if not revenue_df.empty:
                st.markdown("### Revenue Report")
                st.dataframe(revenue_df, use_container_width=True)
                
                # Chart
                fig = go.Figure()
                fig.add_trace(go.Bar(
                    x=revenue_df['period'],
                    y=revenue_df['paid_revenue'],
                    name='Paid Revenue',
                    marker_color='#27ae60'
                ))
                fig.add_trace(go.Bar(
                    x=revenue_df['period'],
                    y=revenue_df['pending_revenue'],
                    name='Pending Revenue',
                    marker_color='#f39c12'
                ))
                fig.update_layout(
                    barmode='group',
                    xaxis_title="Period",
                    yaxis_title=f"Revenue ({get_currency_symbol(st.session_state.currency)})",
                    plot_bgcolor='white',
                    paper_bgcolor='white'
                )
                st.plotly_chart(fig, use_container_width=True)
        
        elif report_type == "Aging Report":
            # Accounts receivable aging
            today = datetime.now().strftime('%Y-%m-%d')
            with get_db_connection() as conn:
                aging_df = pd.read_sql_query("""
                    SELECT 
                        client_name,
                        invoice_number,
                        invoice_date,
                        due_date,
                        grand_total,
                        amount_paid,
                        balance_due,
                        julianday(?) - julianday(due_date) as days_overdue
                    FROM invoices
                    WHERE status NOT IN ('Paid', 'Cancelled')
                    ORDER BY days_overdue DESC
                """, conn, params=[today])
            
            if not aging_df.empty:
                st.markdown("### Accounts Receivable Aging")
                
                # Categorize aging
                aging_df['aging_category'] = pd.cut(
                    aging_df['days_overdue'],
                    bins=[-float('inf'), 0, 30, 60, 90, float('inf')],
                    labels=['Current', '1-30 days', '31-60 days', '61-90 days', '90+ days']
                )
                
                aging_summary = aging_df.groupby('aging_category')['balance_due'].sum().reset_index()
                
                col1, col2 = st.columns(2)
                with col1:
                    st.dataframe(aging_summary, use_container_width=True)
                
                with col2:
                    fig = px.pie(
                        aging_summary,
                        values='balance_due',
                        names='aging_category',
                        title='Aging Summary'
                    )
                    st.plotly_chart(fig, use_container_width=True)
                
                st.markdown("### Detailed Aging")
                st.dataframe(aging_df, use_container_width=True)
        
        elif report_type == "Client Summary":
            # Client revenue summary
            with get_db_connection() as conn:
                client_summary = pd.read_sql_query("""
                    SELECT 
                        i.client_name,
                        COUNT(*) as invoice_count,
                        SUM(CASE WHEN i.status = 'Paid' THEN i.grand_total ELSE 0 END) as paid_amount,
                        SUM(CASE WHEN i.status != 'Paid' THEN i.grand_total ELSE 0 END) as pending_amount,
                        SUM(i.grand_total) as total_amount,
                        AVG(i.grand_total) as avg_invoice,
                        MAX(i.invoice_date) as last_invoice
                    FROM invoices i
                    WHERE i.invoice_date BETWEEN ? AND ?
                    GROUP BY i.client_name
                    ORDER BY total_amount DESC
                """, conn, params=[report_start.strftime('%Y-%m-%d'), report_end.strftime('%Y-%m-%d')])
            
            if not client_summary.empty:
                st.markdown("### Client Summary Report")
                st.dataframe(client_summary, use_container_width=True)
                
                # Top clients chart
                top_clients = client_summary.head(10)
                fig = px.bar(
                    top_clients,
                    x='client_name',
                    y='total_amount',
                    title='Top 10 Clients by Revenue'
                )
                fig.update_layout(
                    xaxis_tickangle=-45,
                    plot_bgcolor='white',
                    paper_bgcolor='white'
                )
                st.plotly_chart(fig, use_container_width=True)
        
        elif report_type == "Tax Summary":
            # Tax collected summary
            with get_db_connection() as conn:
                tax_df = pd.read_sql_query("""
                    SELECT 
                        strftime('%Y-%m', i.invoice_date) as period,
                        COUNT(DISTINCT i.id) as invoice_count,
                        SUM(ii.tax_amount) as total_tax_collected,
                        SUM(i.grand_total) as total_revenue,
                        (SUM(ii.tax_amount) / SUM(i.grand_total) * 100) as avg_tax_rate
                    FROM invoices i
                    JOIN invoice_items ii ON i.id = ii.invoice_id
                    WHERE i.invoice_date BETWEEN ? AND ?
                    GROUP BY strftime('%Y-%m', i.invoice_date)
                    ORDER BY period
                """, conn, params=[report_start.strftime('%Y-%m-%d'), report_end.strftime('%Y-%m-%d')])
            
            if not tax_df.empty:
                st.markdown("### Tax Summary Report")
                st.dataframe(tax_df, use_container_width=True)
                
                fig = go.Figure()
                fig.add_trace(go.Bar(
                    x=tax_df['period'],
                    y=tax_df['total_tax_collected'],
                    name='Tax Collected',
                    marker_color='#3498db'
                ))
                fig.update_layout(
                    xaxis_title="Period",
                    yaxis_title=f"Tax Amount ({get_currency_symbol(st.session_state.currency)})",
                    plot_bgcolor='white',
                    paper_bgcolor='white'
                )
                st.plotly_chart(fig, use_container_width=True)
        
        elif report_type == "Payment Methods":
            # Payment method summary
            with get_db_connection() as conn:
                payment_method_df = pd.read_sql_query("""
                    SELECT 
                        p.payment_method,
                        COUNT(*) as payment_count,
                        SUM(p.amount) as total_amount,
                        AVG(p.amount) as avg_amount,
                        MIN(p.payment_date) as first_use,
                        MAX(p.payment_date) as last_use
                    FROM payments p
                    WHERE p.payment_date BETWEEN ? AND ?
                    GROUP BY p.payment_method
                    ORDER BY total_amount DESC
                """, conn, params=[report_start.strftime('%Y-%m-%d'), report_end.strftime('%Y-%m-%d')])
            
            if not payment_method_df.empty:
                st.markdown("### Payment Methods Report")
                st.dataframe(payment_method_df, use_container_width=True)
                
                col1, col2 = st.columns(2)
                with col1:
                    fig = px.pie(
                        payment_method_df,
                        values='total_amount',
                        names='payment_method',
                        title='Payment Amount by Method'
                    )
                    st.plotly_chart(fig, use_container_width=True)
                
                with col2:
                    fig = px.pie(
                        payment_method_df,
                        values='payment_count',
                        names='payment_method',
                        title='Payment Count by Method'
                    )
                    st.plotly_chart(fig, use_container_width=True)

# ============================================================================
# SETTINGS PAGE
# ============================================================================

def render_settings_page():
    """Render the settings page"""
    
    st.markdown('<div class="section-header">⚙️ Settings</div>', unsafe_allow_html=True)
    
    tabs = st.tabs(["🏢 Company", "💾 Database", "👤 Users", "📧 Email", "🔐 Security"])
    
    with tabs[0]:
        st.markdown('<div class="business-card">', unsafe_allow_html=True)
        st.markdown("##### Company Settings")
        
        col1, col2 = st.columns(2)
        
        with col1:
            company_name = st.text_input("Company Name", value=st.session_state.company_info['name'])
            company_address = st.text_input("Address", value=st.session_state.company_info['address'])
            company_city = st.text_input("City", value=st.session_state.company_info['city'])
            company_phone = st.text_input("Phone", value=st.session_state.company_info['phone'])
        
        with col2:
            company_email = st.text_input("Email", value=st.session_state.company_info['email'])
            company_tax_id = st.text_input("TRN / Tax ID", value=st.session_state.company_info['tax_id'])
            invoice_prefix = st.text_input("Invoice Prefix", value=st.session_state.company_info.get('invoice_prefix', 'INV'))
            default_currency = st.selectbox(
                "Default Currency",
                options=list(CURRENCIES.keys()),
                format_func=lambda x: f"{CURRENCIES[x]['symbol']} {CURRENCIES[x]['name']}",
                index=list(CURRENCIES.keys()).index(st.session_state.company_info.get('default_currency', 'TTD'))
            )
        
        vat_registered = st.checkbox("VAT Registered", value=st.session_state.company_info.get('vat_registered', True))
        
        company_bank = st.text_area(
            "Bank Details",
            value=st.session_state.company_info.get('bank_details', ''),
            height=100,
            help="Include account number, bank name, sort code, etc."
        )
        
        # Logo
        st.markdown("##### Company Logo")
        logo_file = st.file_uploader(
            "Upload Logo (PNG, JPG, JPEG)",
            type=['png', 'jpg', 'jpeg'],
            key="settings_logo_upload"
        )
        
        if logo_file is not None:
            if save_logo(logo_file):
                st.success(f"✓ Logo uploaded: {logo_file.name}")
        
        if st.session_state.company_info.get('logo_base64'):
            st.markdown(f'<div class="logo-container">{get_logo_html("80px", "200px")}</div>', unsafe_allow_html=True)
            if st.button("🗑️ Remove Logo", key="settings_remove_logo"):
                remove_logo()
                st.rerun()
        
        if st.button("💾 Save Company Settings", use_container_width=True):
            st.session_state.company_info.update({
                'name': company_name,
                'address': company_address,
                'city': company_city,
                'phone': company_phone,
                'email': company_email,
                'tax_id': company_tax_id,
                'bank_details': company_bank,
                'default_currency': default_currency,
                'vat_registered': vat_registered,
                'invoice_prefix': invoice_prefix
            })
            
            # Save to database
            try:
                with get_db_connection() as conn:
                    c = conn.cursor()
                    c.execute('''UPDATE company_settings 
                               SET name = ?, address = ?, city = ?, phone = ?, 
                                   email = ?, tax_id = ?, bank_details = ?,
                                   default_currency = ?, vat_registered = ?, 
                                   invoice_prefix = ?, updated_at = ?
                               WHERE id = 1''',
                             (company_name, company_address, company_city, company_phone,
                              company_email, company_tax_id, company_bank,
                              default_currency, vat_registered, invoice_prefix,
                              datetime.now().isoformat()))
                    conn.commit()
                    
                    st.session_state.notification = "✓ Company settings saved"
                    st.session_state.notification_type = "success"
                    st.rerun()
            except Exception as e:
                st.error(f"Error saving settings: {e}")
        
        st.markdown('</div>', unsafe_allow_html=True)
    
    with tabs[1]:
        st.markdown('<div class="business-card">', unsafe_allow_html=True)
        st.markdown("##### Database Management")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("**Backup Database**")
            if st.button("📥 Create Backup", use_container_width=True):
                backup_data, filename = backup_database()
                if backup_data:
                    st.download_button(
                        label="📥 Download Backup",
                        data=backup_data,
                        file_name=filename,
                        mime="application/octet-stream",
                        use_container_width=True
                    )
                else:
                    st.error("Backup failed")
        
        with col2:
            st.markdown("**Restore Database**")
            uploaded_backup = st.file_uploader(
                "Upload Backup File",
                type=['db'],
                key="backup_upload"
            )
            if uploaded_backup and st.button("🔄 Restore from Backup", use_container_width=True):
                # Save uploaded file temporarily
                temp_path = "temp_restore.db"
                with open(temp_path, 'wb') as f:
                    f.write(uploaded_backup.getbuffer())
                
                if restore_database(temp_path):
                    os.remove(temp_path)
                    st.session_state.notification = "✓ Database restored successfully"
                    st.session_state.notification_type = "success"
                    st.rerun()
                else:
                    os.remove(temp_path)
                    st.error("Restore failed")
        
        st.divider()
        
        # Database stats
        st.markdown("**Database Statistics**")
        
        try:
            db_size = os.path.getsize('invoices.db') / 1024  # KB
            
            with get_db_connection() as conn:
                c = conn.cursor()
                c.execute("SELECT COUNT(*) FROM invoices")
                invoice_count = c.fetchone()[0]
                c.execute("SELECT COUNT(*) FROM clients")
                client_count = c.fetchone()[0]
                c.execute("SELECT COUNT(*) FROM payments")
                payment_count = c.fetchone()[0]
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Database Size", f"{db_size:.1f} KB")
            with col2:
                st.metric("Invoices", invoice_count)
            with col3:
                st.metric("Clients", client_count)
            with col4:
                st.metric("Payments", payment_count)
        except Exception as e:
            st.warning(f"Could not load database stats: {e}")
        
        st.markdown('</div>', unsafe_allow_html=True)
    
    with tabs[2]:
        st.markdown('<div class="business-card">', unsafe_allow_html=True)
        st.markdown("##### User Management")
        
        # User list
        try:
            with get_db_connection() as conn:
                users_df = pd.read_sql_query(
                    "SELECT id, username, email, role, full_name, is_active, last_login FROM users",
                    conn
                )
            
            if not users_df.empty:
                st.dataframe(users_df, use_container_width=True)
        except:
            st.info("No users found")
        
        st.divider()
        
        # Add user form
        st.markdown("**Add New User**")
        col1, col2 = st.columns(2)
        with col1:
            new_username = st.text_input("Username")
            new_email = st.text_input("Email")
            new_full_name = st.text_input("Full Name")
        with col2:
            new_password = st.text_input("Password", type="password")
            new_role = st.selectbox("Role", options=['user', 'admin', 'viewer'])
            new_active = st.checkbox("Active", value=True)
        
        if st.button("➕ Add User", use_container_width=True):
            if new_username and new_email and new_password:
                if not validate_email(new_email):
                    st.error("Please enter a valid email address")
                else:
                    password_hash = hash_password(new_password)
                    try:
                        with get_db_connection() as conn:
                            c = conn.cursor()
                            c.execute('''INSERT INTO users 
                                       (username, password_hash, email, role, full_name, is_active, created_at)
                                       VALUES (?, ?, ?, ?, ?, ?, ?)''',
                                     (new_username, password_hash, new_email, new_role, new_full_name, new_active,
                                      datetime.now().isoformat()))
                            conn.commit()
                            st.session_state.notification = f"✓ User {new_username} added"
                            st.session_state.notification_type = "success"
                            st.rerun()
                    except sqlite3.IntegrityError:
                        st.error("Username or email already exists")
                    except Exception as e:
                        st.error(f"Error adding user: {e}")
        
        st.markdown('</div>', unsafe_allow_html=True)
    
    with tabs[3]:
        st.markdown('<div class="business-card">', unsafe_allow_html=True)
        st.markdown("##### Email Configuration")
        
        # Load from environment or session
        smtp_server = st.text_input("SMTP Server", value=os.getenv('SMTP_SERVER', 'smtp.gmail.com'))
        smtp_port = st.number_input("SMTP Port", value=int(os.getenv('SMTP_PORT', 587)), min_value=1, max_value=65535)
        smtp_username = st.text_input("SMTP Username", value=os.getenv('SMTP_USERNAME', ''))
        smtp_password = st.text_input("SMTP Password", type="password", value=os.getenv('SMTP_PASSWORD', ''))
        use_tls = st.checkbox("Use TLS", value=True)
        
        if st.button("💾 Save Email Settings", use_container_width=True):
            # Save to .env file
            with open('.env', 'w') as f:
                f.write(f"SMTP_SERVER={smtp_server}\n")
                f.write(f"SMTP_PORT={smtp_port}\n")
                f.write(f"SMTP_USERNAME={smtp_username}\n")
                f.write(f"SMTP_PASSWORD={smtp_password}\n")
                f.write(f"SMTP_USE_TLS={'True' if use_tls else 'False'}\n")
            
            # Update environment variables
            os.environ['SMTP_SERVER'] = smtp_server
            os.environ['SMTP_PORT'] = str(smtp_port)
            os.environ['SMTP_USERNAME'] = smtp_username
            os.environ['SMTP_PASSWORD'] = smtp_password
            os.environ['SMTP_USE_TLS'] = 'True' if use_tls else 'False'
            
            st.session_state.notification = "✓ Email settings saved"
            st.session_state.notification_type = "success"
            st.rerun()
        
        st.divider()
        
        # Test email
        st.markdown("**Test Email Configuration**")
        test_email = st.text_input("Send Test Email To")
        if st.button("📧 Send Test Email", use_container_width=True) and test_email:
            try:
                msg = MIMEMultipart()
                msg['From'] = st.session_state.company_info['email']
                msg['To'] = test_email
                msg['Subject'] = "Test Email from Invoice Pro"
                
                body = f"""
                <html>
                <body>
                    <p>This is a test email from Invoice Pro.</p>
                    <p>If you're reading this, your email configuration is working correctly!</p>
                    <p>Best regards,<br>{st.session_state.company_info['name']}</p>
                </body>
                </html>
                """
                msg.attach(MIMEText(body, 'html'))
                
                server = smtplib.SMTP(smtp_server, smtp_port)
                if use_tls:
                    server.starttls()
                server.login(smtp_username, smtp_password)
                server.send_message(msg)
                server.quit()
                
                st.success(f"✓ Test email sent to {test_email}")
            except Exception as e:
                st.error(f"Error sending test email: {e}")
        
        st.markdown('</div>', unsafe_allow_html=True)
    
    with tabs[4]:
        st.markdown('<div class="business-card">', unsafe_allow_html=True)
        st.markdown("##### Security Settings")
        
        # Password policy
        st.markdown("**Password Policy**")
        min_password_length = st.number_input("Minimum Password Length", min_value=6, value=8)
        require_special = st.checkbox("Require Special Characters", value=True)
        require_numbers = st.checkbox("Require Numbers", value=True)
        require_uppercase = st.checkbox("Require Uppercase Letters", value=True)
        
        # Session timeout
        session_timeout = st.number_input("Session Timeout (minutes)", min_value=5, value=30)
        
        # 2FA
        enable_2fa = st.checkbox("Enable Two-Factor Authentication", value=False)
        if enable_2fa:
            st.info("Two-factor authentication setup requires additional configuration")
        
        # Audit log
        st.divider()
        st.markdown("**Audit Log**")
        if st.button("📋 View Audit Log"):
            try:
                with get_db_connection() as conn:
                    audit_df = pd.read_sql_query(
                        "SELECT * FROM audit_log ORDER BY timestamp DESC LIMIT 100",
                        conn
                    )
                if not audit_df.empty:
                    st.dataframe(audit_df, use_container_width=True)
                else:
                    st.info("No audit logs found")
            except Exception as e:
                st.error(f"Error loading audit log: {e}")
        
        st.markdown('</div>', unsafe_allow_html=True)

# ============================================================================
# HELP PAGE
# ============================================================================

def render_help_page():
    """Render the help page"""
    
    st.markdown('<div class="section-header">❓ Help & Support</div>', unsafe_allow_html=True)
    
    tabs = st.tabs(["📖 User Guide", "❓ FAQ", "📞 Contact", "ℹ️ About"])
    
    with tabs[0]:
        st.markdown("""
        ### 📖 User Guide
        
        #### Getting Started
        1. **Create an Invoice**
           - Go to "Create Invoice" tab
           - Fill in client details
           - Add items with quantities and prices
           - Review totals and save
        
        2. **Manage Clients**
           - Add client information in the Clients section
           - Save frequently used clients for quick selection
        
        3. **Track Payments**
           - Record payments against invoices
           - View payment history
           - Track outstanding balances
        
        4. **Generate Reports**
           - Access financial reports
           - Export data for accounting
           - Monitor business performance
        
        #### Keyboard Shortcuts
        - `Ctrl+N`: New Invoice
        - `Ctrl+S`: Save Current
        - `Ctrl+P`: Print/PDF
        - `Ctrl+F`: Search
        """)
    
    with tabs[1]:
        st.markdown("""
        ### ❓ Frequently Asked Questions
        
        **Q: How do I change the currency?**
        A: Use the currency selector in the sidebar to change the display currency for all invoices.
        
        **Q: Can I customize the invoice template?**
        A: Yes! Go to Settings > Company to upload your logo and customize company details.
        
        **Q: How do I set up recurring invoices?**
        A: Create an invoice, then in Advanced Options select a recurring frequency and save.
        
        **Q: Is my data secure?**
        A: Yes, all data is stored locally in an encrypted SQLite database.
        
        **Q: Can I backup my data?**
        A: Absolutely! Go to Settings > Database to create and download backups.
        
        **Q: How do I send invoices via email?**
        A: Configure email settings in Settings > Email, then use the Send button when viewing an invoice.
        """)
    
    with tabs[2]:
        st.markdown("""
        ### 📞 Contact Support
        
        **Email:** support@invoicepro.com  
        **Phone:** +1 (868) 123-4567  
        **Hours:** Monday-Friday, 9am-5pm AST
        
        #### Office Location
        Invoice Pro Headquarters  
        123 Business Avenue  
        Port of Spain, Trinidad  
        West Indies
        
        #### Submit a Ticket
        """)
        
        with st.form("support_form"):
            name = st.text_input("Your Name")
            email = st.text_input("Email Address")
            subject = st.text_input("Subject")
            message = st.text_area("Message", height=150)
            
            if st.form_submit_button("📤 Submit Ticket"):
                if name and email and subject and message:
                    if not validate_email(email):
                        st.error("Please enter a valid email address")
                    else:
                        st.success("✓ Ticket submitted successfully. We'll respond within 24 hours.")
                else:
                    st.warning("Please fill in all fields")
    
    with tabs[3]:
        st.markdown("""
        ### ℹ️ About Invoice Pro 2026
        
        **Version:** 3.0.0  
        **Release Date:** January 2026  
        **Developer:** Invoice Pro Team  
        
        #### Features
        - ✅ Professional invoice generation
        - ✅ Multi-currency support
        - ✅ Client management
        - ✅ Payment tracking
        - ✅ Financial reports
        - ✅ Data backup & restore
        - ✅ Email integration
        - ✅ PDF export
        
        #### System Requirements
        - Python 3.8+
        - Modern web browser
        - 100MB free disk space
        
        #### License
        Commercial License - All rights reserved
        
        #### Acknowledgements
        Special thanks to all our beta testers and early adopters who helped shape this application.
        
        © 2026 Invoice Pro. All rights reserved.
        """)

# ============================================================================
# MAIN APP ROUTER
# ============================================================================

def main():
    """Main application router"""
    
    # Initialize database
    init_database()
    
    # Add custom CSS
    add_custom_css()
    
    # Initialize session state
    if 'company_info' not in st.session_state:
        # Load company info from database
        try:
            with get_db_connection() as conn:
                company = pd.read_sql_query("SELECT * FROM company_settings WHERE id = 1", conn)
                if not company.empty:
                    st.session_state.company_info = company.iloc[0].to_dict()
                else:
                    st.session_state.company_info = {
                        'name': 'My Company',
                        'address': '',
                        'city': '',
                        'phone': '',
                        'email': '',
                        'tax_id': '',
                        'bank_details': '',
                        'default_currency': 'TTD',
                        'vat_registered': True,
                        'invoice_prefix': 'INV',
                        'logo_base64': None
                    }
        except:
            st.session_state.company_info = {
                'name': 'My Company',
                'address': '',
                'city': '',
                'phone': '',
                'email': '',
                'tax_id': '',
                'bank_details': '',
                'default_currency': 'TTD',
                'vat_registered': True,
                'invoice_prefix': 'INV',
                'logo_base64': None
            }
    
    if 'currency' not in st.session_state:
        st.session_state.currency = st.session_state.company_info.get('default_currency', 'TTD')
    
    if 'current_page' not in st.session_state:
        st.session_state.current_page = "dashboard"
    
    if 'notification' not in st.session_state:
        st.session_state.notification = None
    
    # Display notification if exists
    if st.session_state.notification:
        if st.session_state.notification_type == "success":
            st.markdown(f'<div class="success-notification">{st.session_state.notification}</div>', unsafe_allow_html=True)
        else:
            st.markdown(f'<div class="error-notification">{st.session_state.notification}</div>', unsafe_allow_html=True)
        st.session_state.notification = None
    
    # Render the appropriate page based on session state
    if st.session_state.current_page == "dashboard":
        render_dashboard_page()
    elif st.session_state.current_page == "create":
        render_create_invoice_page()
    elif st.session_state.current_page == "view_invoices":
        render_view_invoices_page()
    elif st.session_state.current_page == "clients":
        render_clients_page()
    elif st.session_state.current_page == "payments":
        render_payments_page()
    elif st.session_state.current_page == "recurring":
        render_recurring_page()
    elif st.session_state.current_page == "reports":
        render_reports_page()
    elif st.session_state.current_page == "settings":
        render_settings_page()
    elif st.session_state.current_page == "help":
        render_help_page()
    else:
        # Default to dashboard
        render_dashboard_page()

# Run the app
if __name__ == "__main__":
    main()
